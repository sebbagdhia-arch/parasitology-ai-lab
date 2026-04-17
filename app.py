# ╔══════════════════════════════════════════════════════════════════════════════════╗
# ║                  DM SMART LAB AI v7.5 - ULTIMATE SPACE EDITION                 ║
# ║            Diagnostic Parasitologique par Intelligence Artificielle              ║
# ║                                                                                ║
# ║  Développé par:                                                                ║
# ║    • Sebbag Mohamed Dhia Eddine (Expert IA & Conception)                       ║
# ║    • Ben Sghir Mohamed (Expert Laboratoire & Données)                          ║
# ║                                                                                ║
# ║  INFSPM - Ouargla, Algérie 🇩🇿                                                ║
# ╚══════════════════════════════════════════════════════════════════════════════════╝

import streamlit as st
import numpy as np
import pandas as pd
import time
import os
import base64
import hashlib
import random
import json
import io
import sqlite3
import math
from PIL import Image, ImageOps, ImageFilter, ImageEnhance, ImageDraw
from datetime import datetime, timedelta
from fpdf import FPDF
from contextlib import contextmanager

try:
    import bcrypt
    HAS_BCRYPT = True
except ImportError:
    HAS_BCRYPT = False

try:
    import plotly.express as px
    import plotly.graph_objects as go
    HAS_PLOTLY = True
except ImportError:
    HAS_PLOTLY = False

try:
    import qrcode
    HAS_QRCODE = True
except ImportError:
    HAS_QRCODE = False

# ============================================
#  PAGE CONFIG
# ============================================
st.set_page_config(
    page_title="DM Smart Lab AI v7.5",
    page_icon="🧬",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================
#  CONSTANTS
# ============================================
APP_VERSION = "7.5.0"
SECRET_KEY = "dm_smart_lab_2026_ultra_secret"
MAX_LOGIN_ATTEMPTS = 5
LOCKOUT_MINUTES = 10
CONFIDENCE_THRESHOLD = 60
MODEL_INPUT_SIZE = (224, 224)
DB_PATH = "dm_smartlab.db"
MAX_UPLOAD_SIZE = 10 * 1024 * 1024  # 10MB
CACHE_TTL = 3600  # 1 hour

ROLES = {
    "admin": {
        "level": 3,
        "icon": "👑",
        "label": {"fr": "Administrateur Système", "ar": "مدير النظام", "en": "System Administrator"},
        "permissions": ["read", "write", "delete", "admin", "validate", "export", "manage_users", "view_logs"],
        "color": "#ff0040"
    },
    "technician": {
        "level": 2,
        "icon": "🔬",
        "label": {"fr": "Technicien Laboratoire", "ar": "تقني مخبر", "en": "Lab Technician"},
        "permissions": ["read", "write", "validate", "export"],
        "color": "#00f5ff"
    },
    "viewer": {
        "level": 1,
        "icon": "👁️",
        "label": {"fr": "Observateur", "ar": "مراقب", "en": "Viewer"},
        "permissions": ["read", "export"],
        "color": "#00ff88"
    }
}

AUTHORS = {
    "dev1": {
        "name": "Sebbag Mohamed Dhia Eddine",
        "email": "sebbag.dhia@infspm.dz",
        "role": {"fr": "Expert IA & Conception", "ar": "خبير ذكاء اصطناعي و تصميم", "en": "AI & Design Expert"},
        "avatar": "👨‍🔬"
    },
    "dev2": {
        "name": "Ben Sghir Mohamed",
        "email": "bensghir.mohamed@infspm.dz",
        "role": {"fr": "Expert Laboratoire & Données", "ar": "خبير مخبر و بيانات", "en": "Laboratory & Data Expert"},
        "avatar": "👨‍⚕️"
    }
}

INSTITUTION = {
    "name": {
        "fr": "Institut National de Formation Supérieure Paramédicale",
        "ar": "المعهد الوطني للتكوين العالي شبه الطبي",
        "en": "National Institute of Higher Paramedical Training"
    },
    "short": "INFSPM",
    "city": "Ouargla",
    "country": {"fr": "Algérie", "ar": "الجزائر", "en": "Algeria"},
    "year": 2026,
    "website": "www.infspm.dz",
    "phone": "+213 29 70 00 00",
    "address": "Ouargla, Wilaya de Ouargla, Algérie"
}


PROJECT_TITLE = {
    "fr": "Exploration du potentiel de l'intelligence artificielle pour la lecture automatique de l'examen parasitologique à l'état frais",
    "ar": "استكشاف إمكانيات الذكاء الاصطناعي للقراءة الآلية للفحص الطفيلي المباشر",
    "en": "Exploring AI potential for automatic reading of fresh parasitological examination"
}

NEON = {
    "cyan": "#00f5ff", "magenta": "#ff00ff", "green": "#00ff88",
    "orange": "#ff6600", "red": "#ff0040", "blue": "#0066ff",
    "purple": "#9933ff", "yellow": "#ffff00", "pink": "#ff69b4"
}

MICROSCOPE_TYPES = [
    "Microscope Optique", "Microscope Binoculaire", "Microscope Inversé",
    "Microscope à Fluorescence", "Microscope Contraste de Phase",
    "Microscope Fond Noir", "Microscope Numérique", "Microscope Confocal"
]

MAGNIFICATIONS = ["x4", "x10", "x20", "x40", "x60", "x100 (Immersion)"]

PREPARATION_TYPES = [
    "État Frais (Direct)", "Coloration au Lugol", "MIF", "Concentration Ritchie",
    "Kato-Katz", "Coloration MGG", "Coloration Giemsa", "Ziehl-Neelsen Modifié",
    "Coloration Trichrome", "Goutte Épaisse", "Frottis Mince", "Scotch-Test (Graham)",
    "Technique Baermann", "Flottation Willis", "Technique Knott"
]

SAMPLES = {
    "fr": ["Selles", "Sang (Frottis)", "Sang (Goutte épaisse)", "Urines", "LCR",
           "Biopsie Cutanée", "Crachat", "Autre"],
    "ar": ["براز", "دم (لطاخة)", "دم (قطرة سميكة)", "بول", "سائل دماغي شوكي",
           "خزعة جلدية", "بلغم", "أخرى"],
    "en": ["Stool", "Blood (Smear)", "Blood (Thick drop)", "Urine", "CSF",
           "Skin Biopsy", "Sputum", "Other"]
}


# ============================================
#  COMPLETE TRANSLATION SYSTEM
# ============================================
TR = {
    "fr": {
        "app_title": "DM Smart Lab AI",
        "login_title": "Connexion Sécurisée",
        "login_subtitle": "Système d'Authentification Professionnel",
        "username": "Identifiant",
        "password": "Mot de Passe",
        "connect": "SE CONNECTER",
        "logout": "Déconnexion",
        "home": "Accueil",
        "scan": "Scan & Analyse",
        "encyclopedia": "Encyclopédie",
        "dashboard": "Tableau de Bord",
        "quiz": "Quiz Médical",
        "chatbot": "DM Bot",
        "compare": "Comparaison",
        "admin": "Administration",
        "about": "À Propos",
        "greeting_morning": "Bonjour",
        "greeting_afternoon": "Bon après-midi",
        "greeting_evening": "Bonsoir",
        "welcome_btn": "Message de Bienvenue",
        "intro_btn": "Présentation du Système",
        "stop_voice": "Arrêter",
        "patient_info": "Informations du Patient",
        "patient_name": "Nom du Patient",
        "patient_firstname": "Prénom",
        "age": "Âge",
        "sex": "Sexe",
        "male": "Homme",
        "female": "Femme",
        "weight": "Poids (kg)",
        "sample_type": "Type d'Échantillon",
        "lab_info": "Informations du Laboratoire",
        "microscope": "Microscope",
        "magnification": "Grossissement",
        "preparation": "Préparation",
        "technician": "Technicien",
        "notes": "Notes / Observations",
        "image_capture": "Capture Microscopique",
        "take_photo": "Prendre une Photo (Caméra)",
        "upload_file": "Importer un fichier",
        "camera_hint": "Placez l'oculaire du microscope devant la caméra",
        "result": "Résultat",
        "confidence": "Confiance",
        "risk": "Risque",
        "morphology": "Morphologie",
        "description": "Description",
        "advice": "Conseil Médical",
        "extra_tests": "Examens complémentaires suggérés",
        "diagnostic_keys": "Clés Diagnostiques",
        "lifecycle": "Cycle de Vie",
        "all_probabilities": "Toutes les probabilités",
        "download_pdf": "Télécharger PDF",
        "save_db": "Sauvegarder",
        "new_analysis": "Nouvelle Analyse",
        "listen": "Écouter",
        "total_analyses": "Total Analyses",
        "reliable": "Fiables",
        "to_verify": "À Vérifier",
        "most_frequent": "Plus Fréquent",
        "avg_confidence": "Confiance Moy.",
        "parasite_distribution": "Distribution des Parasites",
        "confidence_levels": "Niveaux de Confiance",
        "trends": "Tendances (30 jours)",
        "history": "Historique Complet",
        "validate": "Valider",
        "export_csv": "CSV",
        "export_json": "JSON",
        "start_quiz": "Démarrer le Quiz",
        "next_question": "Question Suivante",
        "restart": "Recommencer",
        "leaderboard": "Classement",
        "score_excellent": "Excellent ! Vous maîtrisez la parasitologie !",
        "score_good": "Bien joué ! Continuez à apprendre !",
        "score_average": "Pas mal ! Révisez encore un peu !",
        "score_low": "Courage ! La parasitologie s'apprend avec la pratique !",
        "search": "Rechercher...",
        "no_data": "Aucune donnée disponible",
        "no_results": "Aucun résultat trouvé",
        "language": "Langue",
        "daily_tip": "Conseil du Jour",
        "users_mgmt": "Gestion des Utilisateurs",
        "activity_log": "Journal d'Activité",
        "system_info": "Système",
        "create_user": "Créer un Utilisateur",
        "change_pwd": "Changer le Mot de Passe",
        "image1": "Image 1 (Avant)",
        "image2": "Image 2 (Après)",
        "compare_btn": "Comparer les Images",
        "similarity": "Similarité",
        "filter_comparison": "Comparaison des Filtres",
        "pixel_diff": "Différence Pixel à Pixel",
        "name_required": "Le nom du patient est obligatoire !",
        "saved_ok": "Résultat sauvegardé !",
        "demo_mode": "Mode démonstration (aucun modèle IA chargé)",
        "low_conf_warn": "Confiance faible. Vérification manuelle recommandée !",
        "voice_welcome": "Bienvenue dans DM Smart Lab AI ! Nous sommes ravis de vous accueillir dans ce système d'intelligence artificielle dédié au diagnostic parasitologique. Ce système a été conçu pour assister les professionnels de santé dans l'identification rapide et précise des parasites.",
        "voice_intro": "Je suis DM Smart Lab AI, version 7 point 5, système de diagnostic parasitologique par intelligence artificielle. J'ai été développé par deux techniciens supérieurs de l'Institut National de Formation Supérieure Paramédicale de Ouargla. Sebbag Mohamed Dhia Eddine, expert en intelligence artificielle et conception, et Ben Sghir Mohamed, expert en laboratoire et données. Ensemble, nous repoussons les limites de la parasitologie moderne !",
        "quick_overview": "Aperçu Rapide",
        "where_science": "Où la Science Rencontre l'Intelligence",
        "system_desc": "Système de diagnostic parasitologique assisté par IA",
        "dev_team": "Équipe de Développement",
        "institution": "Établissement",
        "technologies": "Technologies Utilisées",
        "chat_welcome": "Bonjour ! Je suis **DM Bot**, votre assistant parasitologique intelligent.\n\nJe peux vous aider avec :\n- **Parasites** : Amoeba, Giardia, Plasmodium, Leishmania, Trypanosoma, Schistosoma, Ascaris, Taenia, Toxoplasma, Oxyure, Cryptosporidium...\n- **Techniques** : Microscopie, Colorations, Concentration, EPS...\n- **Traitements** : Protocoles thérapeutiques\n- **Cas cliniques** : Diagnostic différentiel\n\nTapez un mot-clé pour commencer !",
        "chat_placeholder": "Posez votre question sur les parasites...",
        "chat_not_found": "Je n'ai pas trouvé de réponse exacte. Essayez avec un mot-clé comme : **amoeba**, **giardia**, **plasmodium**, **microscope**, **coloration**, **traitement**, **concentration**, **toxoplasma**, **ascaris**, **taenia**, **oxyure** ou tapez **aide** pour voir tout ce que je connais !",
        "clear_chat": "Effacer le chat",
        "quick_questions": "Questions rapides :",
    },
    "ar": {
        "app_title": "مختبر DM الذكي",
        "login_title": "تسجيل الدخول الآمن",
        "login_subtitle": "نظام المصادقة المهني",
        "username": "اسم المستخدم",
        "password": "كلمة المرور",
        "connect": "تسجيل الدخول",
        "logout": "تسجيل الخروج",
        "home": "الرئيسية",
        "scan": "مسح و تحليل",
        "encyclopedia": "الموسوعة",
        "dashboard": "لوحة التحكم",
        "quiz": "اختبار طبي",
        "chatbot": "DM بوت",
        "compare": "المقارنة",
        "admin": "الإدارة",
        "about": "حول المشروع",
        "greeting_morning": "صباح الخير",
        "greeting_afternoon": "مساء الخير",
        "greeting_evening": "مساء الخير",
        "welcome_btn": "رسالة ترحيبية",
        "intro_btn": "تقديم النظام",
        "stop_voice": "إيقاف",
        "patient_info": "معلومات المريض",
        "patient_name": "اسم المريض",
        "patient_firstname": "الاسم الأول",
        "age": "العمر",
        "sex": "الجنس",
        "male": "ذكر",
        "female": "أنثى",
        "weight": "الوزن (كغ)",
        "sample_type": "نوع العينة",
        "lab_info": "معلومات المخبر",
        "microscope": "المجهر",
        "magnification": "التكبير",
        "preparation": "نوع التحضير",
        "technician": "التقني",
        "notes": "ملاحظات",
        "image_capture": "التقاط مجهري",
        "take_photo": "التقاط صورة (الكاميرا)",
        "upload_file": "استيراد ملف",
        "camera_hint": "ضع عدسة المجهر أمام الكاميرا",
        "result": "النتيجة",
        "confidence": "نسبة الثقة",
        "risk": "مستوى الخطر",
        "morphology": "المورفولوجيا",
        "description": "الوصف",
        "advice": "النصيحة الطبية",
        "extra_tests": "فحوصات إضافية مقترحة",
        "diagnostic_keys": "مفاتيح التشخيص",
        "lifecycle": "دورة الحياة",
        "all_probabilities": "جميع الاحتمالات",
        "download_pdf": "تحميل PDF",
        "save_db": "حفظ",
        "new_analysis": "تحليل جديد",
        "listen": "استماع",
        "total_analyses": "مجموع التحاليل",
        "reliable": "موثوقة",
        "to_verify": "للتحقق",
        "most_frequent": "الأكثر شيوعاً",
        "avg_confidence": "متوسط الثقة",
        "parasite_distribution": "توزيع الطفيليات",
        "confidence_levels": "مستويات الثقة",
        "trends": "الاتجاهات (30 يوم)",
        "history": "السجل الكامل",
        "validate": "مصادقة",
        "export_csv": "CSV",
        "export_json": "JSON",
        "start_quiz": "بدء الاختبار",
        "next_question": "السؤال التالي",
        "restart": "إعادة",
        "leaderboard": "الترتيب",
        "score_excellent": "ممتاز ! أنت خبير في علم الطفيليات !",
        "score_good": "أحسنت ! واصل التعلم !",
        "score_average": "لا بأس ! راجع قليلاً !",
        "score_low": "شجاعة ! علم الطفيليات يُتعلم بالممارسة !",
        "search": "بحث...",
        "no_data": "لا توجد بيانات",
        "no_results": "لا توجد نتائج",
        "language": "اللغة",
        "daily_tip": "نصيحة اليوم",
        "users_mgmt": "إدارة المستخدمين",
        "activity_log": "سجل النشاط",
        "system_info": "النظام",
        "create_user": "إنشاء مستخدم",
        "change_pwd": "تغيير كلمة المرور",
        "image1": "الصورة 1 (قبل)",
        "image2": "الصورة 2 (بعد)",
        "compare_btn": "مقارنة الصور",
        "similarity": "التشابه",
        "filter_comparison": "مقارنة الفلاتر",
        "pixel_diff": "الفرق بكسل ببكسل",
        "name_required": "اسم المريض مطلوب !",
        "saved_ok": "تم الحفظ بنجاح !",
        "demo_mode": "وضع تجريبي (لا يوجد نموذج ذكاء اصطناعي)",
        "low_conf_warn": "ثقة منخفضة. يُنصح بالتحقق اليدوي !",
        "voice_welcome": "مرحباً بكم في مختبر DM الذكي. نحن سعداء باستقبالكم في هذا النظام المخصص للتشخيص الطفيلي بالذكاء الاصطناعي.",
        "voice_intro": "أنا مختبر DM الذكي، النسخة 7.5، نظام تشخيص طفيلي بالذكاء الاصطناعي. تم تطويري من طرف تقنيين ساميين من المعهد الوطني للتكوين العالي شبه الطبي بورقلة.",
        "quick_overview": "نظرة سريعة",
        "where_science": "حيث يلتقي العلم بالذكاء",
        "system_desc": "نظام تشخيص طفيلي بالذكاء الاصطناعي",
        "dev_team": "فريق التطوير",
        "institution": "المؤسسة",
        "technologies": "التقنيات المستخدمة",
        "chat_welcome": "مرحباً! أنا **DM Bot**، مساعدك الذكي في علم الطفيليات.\n\nأستطيع مساعدتك في:\n- **الطفيليات**: الأميبا، الجيارديا، البلازموديوم، الليشمانيا...\n- **التقنيات**: المجهر، التلوينات، التركيز...\n- **العلاجات**: البروتوكولات العلاجية\n\naكتب كلمة مفتاحية للبدء!",
        "chat_placeholder": "اطرح سؤالك عن الطفيليات...",
        "chat_not_found": "لم أجد إجابة دقيقة. جرب كلمة مفتاحية مثل: **amoeba**، **giardia**، **plasmodium**، أو اكتب **aide** لرؤية كل ما أعرفه!",
        "clear_chat": "مسح المحادثة",
        "quick_questions": "أسئلة سريعة:",
    },
    "en": {
        "app_title": "DM Smart Lab AI",
        "login_title": "Secure Login",
        "login_subtitle": "Professional Authentication System",
        "username": "Username",
        "password": "Password",
        "connect": "LOG IN",
        "logout": "Logout",
        "home": "Home",
        "scan": "Scan & Analysis",
        "encyclopedia": "Encyclopedia",
        "dashboard": "Dashboard",
        "quiz": "Medical Quiz",
        "chatbot": "DM Bot",
        "compare": "Comparison",
        "admin": "Administration",
        "about": "About",
        "greeting_morning": "Good morning",
        "greeting_afternoon": "Good afternoon",
        "greeting_evening": "Good evening",
        "welcome_btn": "Welcome Message",
        "intro_btn": "System Introduction",
        "stop_voice": "Stop",
        "patient_info": "Patient Information",
        "patient_name": "Patient Name",
        "patient_firstname": "First Name",
        "age": "Age",
        "sex": "Sex",
        "male": "Male",
        "female": "Female",
        "weight": "Weight (kg)",
        "sample_type": "Sample Type",
        "lab_info": "Laboratory Information",
        "microscope": "Microscope",
        "magnification": "Magnification",
        "preparation": "Preparation",
        "technician": "Technician",
        "notes": "Notes / Observations",
        "image_capture": "Microscopic Capture",
        "take_photo": "Take a Photo (Camera)",
        "upload_file": "Upload a file",
        "camera_hint": "Place the microscope eyepiece in front of the camera",
        "result": "Result",
        "confidence": "Confidence",
        "risk": "Risk Level",
        "morphology": "Morphology",
        "description": "Description",
        "advice": "Medical Advice",
        "extra_tests": "Suggested Additional Tests",
        "diagnostic_keys": "Diagnostic Keys",
        "lifecycle": "Life Cycle",
        "all_probabilities": "All Probabilities",
        "download_pdf": "Download PDF",
        "save_db": "Save",
        "new_analysis": "New Analysis",
        "listen": "Listen",
        "total_analyses": "Total Analyses",
        "reliable": "Reliable",
        "to_verify": "To Verify",
        "most_frequent": "Most Frequent",
        "avg_confidence": "Avg. Confidence",
        "parasite_distribution": "Parasite Distribution",
        "confidence_levels": "Confidence Levels",
        "trends": "Trends (30 days)",
        "history": "Complete History",
        "validate": "Validate",
        "export_csv": "CSV",
        "export_json": "JSON",
        "start_quiz": "Start Quiz",
        "next_question": "Next Question",
        "restart": "Restart",
        "leaderboard": "Leaderboard",
        "score_excellent": "Excellent! You master parasitology!",
        "score_good": "Well done! Keep learning!",
        "score_average": "Not bad! Review a bit more!",
        "score_low": "Courage! Parasitology is learned through practice!",
        "search": "Search...",
        "no_data": "No data available",
        "no_results": "No results found",
        "language": "Language",
        "daily_tip": "Daily Tip",
        "users_mgmt": "Users Management",
        "activity_log": "Activity Log",
        "system_info": "System",
        "create_user": "Create User",
        "change_pwd": "Change Password",
        "image1": "Image 1 (Before)",
        "image2": "Image 2 (After)",
        "compare_btn": "Compare Images",
        "similarity": "Similarity",
        "filter_comparison": "Filter Comparison",
        "pixel_diff": "Pixel-by-Pixel Difference",
        "name_required": "Patient name is required!",
        "saved_ok": "Result saved!",
        "demo_mode": "Demo mode (no AI model loaded)",
        "low_conf_warn": "Low confidence. Manual verification recommended!",
        "voice_welcome": "Welcome to DM Smart Lab AI! We are delighted to have you in this artificial intelligence system dedicated to parasitological diagnosis.",
        "voice_intro": "I am DM Smart Lab AI, version 7 point 5, a parasitological diagnosis system powered by artificial intelligence. Developed at INFSPM Ouargla, Algeria.",
        "quick_overview": "Quick Overview",
        "where_science": "Where Science Meets Intelligence",
        "system_desc": "AI-powered parasitological diagnosis system",
        "dev_team": "Development Team",
        "institution": "Institution",
        "technologies": "Technologies Used",
        "chat_welcome": "Hello! I'm **DM Bot**, your intelligent parasitology assistant.\n\nI can help you with:\n- **Parasites**: Amoeba, Giardia, Plasmodium, Leishmania...\n- **Techniques**: Microscopy, Staining, Concentration...\n- **Treatments**: Therapeutic protocols\n\nType a keyword to start!",
        "chat_placeholder": "Ask your question about parasites...",
        "chat_not_found": "I couldn't find an exact answer. Try a keyword like: **amoeba**, **giardia**, **plasmodium**, **microscope**, or type **help** to see everything I know!",
        "clear_chat": "Clear chat",
        "quick_questions": "Quick questions:",
    }
}


def t(key):
    lang = st.session_state.get("lang", "fr")
    return TR.get(lang, TR["fr"]).get(key, TR["fr"].get(key, key))


def tl(d):
    if not isinstance(d, dict):
        return str(d)
    lang = st.session_state.get("lang", "fr")
    return d.get(lang, d.get("fr", str(d)))


def get_greeting():
    h = datetime.now().hour
    if h < 12:
        return t("greeting_morning")
    elif h < 18:
        return t("greeting_afternoon")
    return t("greeting_evening")


# ============================================
#  DATABASE
# ============================================
DB_PATH = "dm_smartlab.db"


@contextmanager
def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def init_database():
    with get_db() as c:
        c.execute("""CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL, full_name TEXT NOT NULL,
            role TEXT DEFAULT 'viewer', speciality TEXT DEFAULT 'Laboratoire',
            is_active INTEGER DEFAULT 1, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            last_login TIMESTAMP, login_count INTEGER DEFAULT 0,
            failed_attempts INTEGER DEFAULT 0, locked_until TIMESTAMP)""")
        c.execute("""CREATE TABLE IF NOT EXISTS analyses (
            id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL,
            patient_name TEXT NOT NULL, patient_firstname TEXT, patient_age INTEGER,
            patient_sex TEXT, patient_weight REAL, sample_type TEXT,
            microscope_type TEXT, magnification TEXT, preparation_type TEXT,
            technician1 TEXT, technician2 TEXT, notes TEXT,
            parasite_detected TEXT NOT NULL, confidence REAL NOT NULL,
            risk_level TEXT, is_reliable INTEGER, all_predictions TEXT,
            image_hash TEXT, is_demo INTEGER DEFAULT 0,
            analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            validated INTEGER DEFAULT 0, validated_by TEXT, validation_date TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id))""")
        c.execute("""CREATE TABLE IF NOT EXISTS activity_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER, username TEXT,
            action TEXT NOT NULL, details TEXT, timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
        c.execute("""CREATE TABLE IF NOT EXISTS quiz_scores (
            id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER, username TEXT,
            score INTEGER NOT NULL, total_questions INTEGER NOT NULL,
            percentage REAL NOT NULL, category TEXT,
            timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
        _make_defaults(c)


def _hp(pw):
    if HAS_BCRYPT:
        return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()
    return hashlib.sha256((pw + SECRET_KEY).encode()).hexdigest()


def _vp(pw, h):
    if HAS_BCRYPT:
        try:
            return bcrypt.checkpw(pw.encode(), h.encode())
        except Exception:
            pass
    return hashlib.sha256((pw + SECRET_KEY).encode()).hexdigest() == h


def _make_defaults(c):
    for u, p, n, r, s in [
        ("admin", "admin2026", "Administrateur Systeme", "admin", "Administration"),
        ("dhia", "dhia2026", "Sebbag Mohamed Dhia Eddine", "admin", "IA & Conception"),
        ("mohamed", "mohamed2026", "Ben Sghir Mohamed", "technician", "Laboratoire"),
        ("demo", "demo123", "Utilisateur Demo", "viewer", "Demonstration"),
        ("tech1", "tech2026", "Technicien Labo 1", "technician", "Parasitologie"),
    ]:
        if not c.execute("SELECT 1 FROM users WHERE username=?", (u,)).fetchone():
            c.execute("INSERT INTO users(username,password_hash,full_name,role,speciality) VALUES(?,?,?,?,?)",
                      (u, _hp(p), n, r, s))


def db_login(u, p):
    with get_db() as c:
        row = c.execute("SELECT * FROM users WHERE username=? AND is_active=1", (u,)).fetchone()
        if not row:
            return None
        if row["locked_until"]:
            try:
                if datetime.now() < datetime.fromisoformat(row["locked_until"]):
                    return {"error": "locked"}
                c.execute("UPDATE users SET failed_attempts=0,locked_until=NULL WHERE id=?", (row["id"],))
            except Exception:
                pass
        if _vp(p, row["password_hash"]):
            c.execute("UPDATE users SET last_login=?,login_count=login_count+1,failed_attempts=0,locked_until=NULL WHERE id=?",
                      (datetime.now().isoformat(), row["id"]))
            return dict(row)
        na = row["failed_attempts"] + 1
        lk = (datetime.now() + timedelta(minutes=LOCKOUT_MINUTES)).isoformat() if na >= MAX_LOGIN_ATTEMPTS else None
        c.execute("UPDATE users SET failed_attempts=?,locked_until=? WHERE id=?", (na, lk, row["id"]))
        return {"error": "wrong", "left": MAX_LOGIN_ATTEMPTS - na}


def db_create_user(u, p, n, r="viewer", s=""):
    with get_db() as c:
        if c.execute("SELECT 1 FROM users WHERE username=?", (u,)).fetchone():
            return False
        c.execute("INSERT INTO users(username,password_hash,full_name,role,speciality) VALUES(?,?,?,?,?)",
                  (u, _hp(p), n, r, s))
        return True


def db_users():
    with get_db() as c:
        return [dict(r) for r in c.execute(
            "SELECT id,username,full_name,role,is_active,last_login,login_count,speciality FROM users").fetchall()]


def db_toggle(uid, active):
    with get_db() as c:
        c.execute("UPDATE users SET is_active=? WHERE id=?", (1 if active else 0, uid))


def db_chpw(uid, pw):
    with get_db() as c:
        c.execute("UPDATE users SET password_hash=? WHERE id=?", (_hp(pw), uid))


def db_save_analysis(uid, d):
    with get_db() as c:
        c.execute("""INSERT INTO analyses(user_id,patient_name,patient_firstname,patient_age,patient_sex,
            patient_weight,sample_type,microscope_type,magnification,preparation_type,technician1,technician2,
            notes,parasite_detected,confidence,risk_level,is_reliable,all_predictions,image_hash,is_demo)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                  (uid, d.get("pn", ""), d.get("pf", ""), d.get("pa", 0), d.get("ps", ""),
                   d.get("pw", 0), d.get("st", ""), d.get("mt", ""), d.get("mg", ""),
                   d.get("pt", ""), d.get("t1", ""), d.get("t2", ""), d.get("nt", ""),
                   d.get("label", "Negative"), d.get("conf", 0), d.get("risk", "none"),
                   d.get("rel", 0), json.dumps(d.get("preds", {})), d.get("hash", ""),
                   d.get("demo", 0)))
        return c.execute("SELECT last_insert_rowid()").fetchone()[0]


def db_analyses(uid=None, lim=500):
    with get_db() as c:
        if uid:
            q = "SELECT a.*,u.full_name as analyst FROM analyses a JOIN users u ON a.user_id=u.id WHERE a.user_id=? ORDER BY a.analysis_date DESC LIMIT ?"
            rows = c.execute(q, (uid, lim)).fetchall()
        else:
            q = "SELECT a.*,u.full_name as analyst FROM analyses a JOIN users u ON a.user_id=u.id ORDER BY a.analysis_date DESC LIMIT ?"
            rows = c.execute(q, (lim,)).fetchall()
        return [dict(r) for r in rows]


def db_stats(uid=None):
    with get_db() as c:
        w = "WHERE user_id=?" if uid else ""
        p = (uid,) if uid else ()
        tot = c.execute(f"SELECT COUNT(*) FROM analyses {w}", p).fetchone()[0]
        if uid:
            rel = c.execute("SELECT COUNT(*) FROM analyses WHERE user_id=? AND is_reliable=1", (uid,)).fetchone()[0]
        else:
            rel = c.execute("SELECT COUNT(*) FROM analyses WHERE is_reliable=1").fetchone()[0]
        para = c.execute(f"SELECT parasite_detected,COUNT(*) as n FROM analyses {w} GROUP BY parasite_detected ORDER BY n DESC", p).fetchall()
        avg = c.execute(f"SELECT AVG(confidence) FROM analyses {w}", p).fetchone()[0] or 0
        return {"total": tot, "reliable": rel, "verify": tot - rel,
                "parasites": [dict(x) for x in para], "avg": round(avg, 1),
                "top": para[0]["parasite_detected"] if para else "N/A"}


def db_trends(days=30):
    with get_db() as c:
        return [dict(r) for r in c.execute("""
            SELECT DATE(analysis_date) as day,parasite_detected,COUNT(*) as count,AVG(confidence) as avg_conf
            FROM analyses WHERE analysis_date>=date('now',?) GROUP BY day,parasite_detected ORDER BY day
        """, (f"-{days} days",)).fetchall()]


def db_log(uid, uname, action, details=""):
    try:
        with get_db() as c:
            c.execute("INSERT INTO activity_log(user_id,username,action,details) VALUES(?,?,?,?)",
                      (uid, uname, action, details))
    except Exception:
        pass


def db_logs(lim=300):
    with get_db() as c:
        return [dict(r) for r in c.execute("SELECT * FROM activity_log ORDER BY timestamp DESC LIMIT ?", (lim,)).fetchall()]


def db_quiz_save(uid, un, sc, tot, pct, cat="general"):
    with get_db() as c:
        c.execute("INSERT INTO quiz_scores(user_id,username,score,total_questions,percentage,category) VALUES(?,?,?,?,?,?)",
                  (uid, un, sc, tot, pct, cat))


def db_leaderboard(lim=20):
    with get_db() as c:
        return [dict(r) for r in c.execute(
            "SELECT username,score,total_questions,percentage,timestamp FROM quiz_scores ORDER BY percentage DESC,timestamp ASC LIMIT ?",
            (lim,)).fetchall()]


def db_validate(aid, who):
    with get_db() as c:
        c.execute("UPDATE analyses SET validated=1,validated_by=?,validation_date=? WHERE id=?",
                  (who, datetime.now().isoformat(), aid))


init_database()


# ============================================
#  PARASITE DATABASE
# ============================================
PARASITE_DB = {
    "Amoeba (E. histolytica)": {
        "sci": "Entamoeba histolytica",
        "morph": {"fr": "Kyste spherique (10-15um) a 4 noyaux, corps chromatoide en cigare. Trophozoite (20-40um) avec pseudopodes digitiformes et hematies phagocytees.",
                  "ar": "كيس كروي (10-15 ميكرومتر) بـ 4 أنوية، جسم كروماتيني على شكل سيجار. الطور النشط (20-40 ميكرومتر) مع أقدام كاذبة وكريات حمراء مبتلعة.",
                  "en": "Spherical cyst (10-15um) with 4 nuclei, cigar-shaped chromatoid body. Trophozoite (20-40um) with pseudopods and phagocytosed RBCs."},
        "desc": {"fr": "Protozoaire responsable de l'amibiase intestinale et extra-intestinale. Transmission feco-orale.",
                 "ar": "طفيلي أولي مسؤول عن الأميبيا المعوية والخارج معوية. الانتقال عبر الفم-البراز.",
                 "en": "Protozoan causing intestinal and extra-intestinal amebiasis. Fecal-oral transmission."},
        "funny": {"fr": "Le ninja des intestins ! Il mange des globules rouges au petit-dejeuner !",
                  "ar": "نينجا الأمعاء! يأكل كريات الدم الحمراء في الفطور!",
                  "en": "The intestinal ninja! Eats red blood cells for breakfast!"},
        "risk": "high",
        "risk_d": {"fr": "Eleve", "ar": "مرتفع", "en": "High"},
        "advice": {"fr": "Metronidazole 500mg x3/j (7-10j) + Amoebicide de contact (Intetrix). Controle EPS J15/J30.",
                   "ar": "ميترونيدازول 500 ملغ 3 مرات يوميا (7-10 أيام) + أميبيسيد تلامسي. مراقبة بعد 15 و 30 يوم.",
                   "en": "Metronidazole 500mg x3/d (7-10d) + Contact amoebicide (Intetrix). Follow-up D15/D30."},
        "tests": ["Serologie amibienne", "Echographie hepatique", "NFS+CRP", "PCR Entamoeba", "Scanner abdominal"],
        "color": "#ff0040", "icon": "🔴",
        "cycle": {"fr": "Kyste ingere -> Excystation -> Trophozoite -> Invasion tissulaire -> Enkystement -> Emission",
                  "ar": "كيس مبتلع ثم انفكاس ثم طور نشط ثم غزو أنسجة ثم تكيس ثم إخراج",
                  "en": "Ingested cyst -> Excystation -> Trophozoite -> Tissue invasion -> Encystation -> Emission"},
        "keys": {"fr": "E. histolytica vs E. dispar: seule histolytica phagocyte les hematies\nKyste 4 noyaux (vs E. coli 8)\nCorps chromatoides en cigare\nMobilite directionnelle",
                 "ar": "E. histolytica مقابل E. dispar: فقط histolytica تبتلع الكريات\nكيس 4 أنوية مقابل 8 لـ E. coli\nأجسام كروماتينية سيجارية\nحركة اتجاهية",
                 "en": "E. histolytica vs E. dispar: only histolytica phagocytoses RBCs\n4 nuclei cyst (vs E. coli 8)\nCigar chromatoid bodies\nDirectional motility"}
    },
    "Giardia": {
        "sci": "Giardia lamblia (intestinalis)",
        "morph": {"fr": "Trophozoite piriforme en cerf-volant (12-15um), 2 noyaux (face de hibou), disque adhesif, 4 paires de flagelles. Kyste ovoide (8-12um) a 4 noyaux.",
                  "ar": "الطور النشط كمثري شكل طائرة ورقية (12-15 ميكرومتر)، نواتان (وجه البومة)، قرص لاصق، 4 أزواج أسواط. كيس بيضاوي (8-12 ميكرومتر) بـ 4 أنوية.",
                  "en": "Pear-shaped kite trophozoite (12-15um), 2 nuclei (owl face), adhesive disk, 4 flagella pairs. Ovoid cyst (8-12um) with 4 nuclei."},
        "desc": {"fr": "Flagelle du duodenum. Diarrhee graisseuse chronique, malabsorption. Transmission hydrique.",
                 "ar": "سوطي الاثني عشر. إسهال دهني مزمن، سوء امتصاص. انتقال عبر الماء.",
                 "en": "Duodenal flagellate. Chronic greasy diarrhea, malabsorption. Waterborne."},
        "funny": {"fr": "Il te fixe avec ses lunettes de soleil ! Un touriste qui refuse de partir !",
                  "ar": "ينظر إليك بنظارته الشمسية! سائح يرفض المغادرة!",
                  "en": "It stares at you with sunglasses! A tourist who refuses to leave!"},
        "risk": "medium", "risk_d": {"fr": "Moyen", "ar": "متوسط", "en": "Medium"},
        "advice": {"fr": "Metronidazole 250mg x3/j (5j) OU Tinidazole 2g dose unique.",
                   "ar": "ميترونيدازول 250 ملغ 3 مرات يوميا (5 أيام) أو تينيدازول 2 غرام جرعة واحدة.",
                   "en": "Metronidazole 250mg x3/d (5d) OR Tinidazole 2g single dose."},
        "tests": ["Ag Giardia ELISA", "Test malabsorption", "EPS x3", "PCR Giardia"],
        "color": "#ff9500", "icon": "🟠",
        "cycle": {"fr": "Kyste ingere -> Excystation duodenale -> Trophozoite -> Adhesion -> Multiplication -> Enkystement",
                  "ar": "كيس مبتلع ثم انفكاس ثم طور نشط ثم التصاق ثم تكاثر ثم تكيس",
                  "en": "Ingested cyst -> Duodenal excystation -> Trophozoite -> Adhesion -> Multiplication -> Encystation"},
        "keys": {"fr": "Forme cerf-volant pathognomonique\n2 noyaux = face de hibou\nDisque adhesif au Lugol\nMobilite feuille morte",
                 "ar": "شكل طائرة ورقية مميز\nنواتان = وجه البومة\nالقرص اللاصق باللوغول\nحركة ورقة ميتة",
                 "en": "Pathognomonic kite shape\n2 nuclei = owl face\nAdhesive disk on Lugol\nFalling leaf motility"}
    },
    "Leishmania": {
        "sci": "Leishmania infantum / major / tropica",
        "morph": {"fr": "Amastigotes ovoides (2-5um) intracellulaires dans macrophages. Noyau + kinetoplaste (MGG).",
                  "ar": "أماستيغوت بيضاوية (2-5 ميكرومتر) داخل البلاعم. نواة + كينيتوبلاست.",
                  "en": "Ovoid amastigotes (2-5um) intracellular in macrophages. Nucleus + kinetoplast (MGG)."},
        "desc": {"fr": "Transmis par phlebotome. Cutanee ou viscerale. Algerie: L. infantum (nord), L. major (sud).",
                 "ar": "ينتقل عبر ذبابة الرمل. جلدية أو حشوية. الجزائر: L. infantum (شمال)، L. major (جنوب).",
                 "en": "Sandfly-transmitted. Cutaneous or visceral. Algeria: L. infantum (north), L. major (south)."},
        "funny": {"fr": "Petit mais costaud ! Il squatte les macrophages !",
                  "ar": "صغير لكن قوي! يحتل البلاعم!",
                  "en": "Small but tough! Squats in macrophages!"},
        "risk": "high", "risk_d": {"fr": "Eleve", "ar": "مرتفع", "en": "High"},
        "advice": {"fr": "Cutanee: Glucantime IM. Viscerale: Amphotericine B liposomale. MDO en Algerie.",
                   "ar": "جلدية: غلوكانتيم عضليا. حشوية: أمفوتيريسين ب. تبليغ إجباري.",
                   "en": "Cutaneous: Glucantime IM. Visceral: Liposomal Amphotericin B. Notifiable."},
        "tests": ["IDR Montenegro", "Serologie", "Ponction medullaire", "Biopsie+MGG", "PCR Leishmania", "NFS"],
        "color": "#ff0040", "icon": "🔴",
        "cycle": {"fr": "Piqure phlebotome -> Promastigotes -> Phagocytose -> Amastigotes -> Multiplication -> Lyse",
                  "ar": "لدغة ذبابة رمل ثم بروماستيغوت ثم بلعمة ثم أماستيغوت ثم تكاثر ثم تحلل",
                  "en": "Sandfly bite -> Promastigotes -> Phagocytosis -> Amastigotes -> Multiplication -> Lysis"},
        "keys": {"fr": "Amastigotes 2-5um intracellulaires\nNoyau + kinetoplaste MGG\nCulture NNN\nPCR = gold standard",
                 "ar": "أماستيغوت 2-5 ميكرومتر داخل خلوية\nنواة + كينيتوبلاست\nزراعة NNN\nPCR المعيار الذهبي",
                 "en": "2-5um intracellular amastigotes\nNucleus+kinetoplast MGG\nNNN culture\nPCR=gold standard"}
    },
    "Plasmodium": {
        "sci": "Plasmodium falciparum / vivax / ovale / malariae",
        "morph": {"fr": "P. falciparum: anneau bague a chaton, gametocytes en banane. P. vivax: trophozoite amiboide, granulations Schuffner.",
                  "ar": "P. falciparum: حلقة خاتم، خلايا جنسية موزية. P. vivax: طور نشط أميبي، حبيبات شوفنر.",
                  "en": "P. falciparum: signet ring, banana gametocytes. P. vivax: amoeboid trophozoite, Schuffner dots."},
        "desc": {"fr": "URGENCE MEDICALE ! Paludisme. P. falciparum = le plus mortel. Anophele femelle.",
                 "ar": "حالة طوارئ طبية! ملاريا. P. falciparum = الأكثر فتكا. أنثى الأنوفيل.",
                 "en": "MEDICAL EMERGENCY! Malaria. P. falciparum = most lethal. Female Anopheles."},
        "funny": {"fr": "Il demande le mariage a tes globules ! Gametocytes en banane = le clown du microscope !",
                  "ar": "يطلب الزواج من كرياتك! خلايا جنسية موزية = مهرج المجهر!",
                  "en": "Proposes to your blood cells! Banana gametocytes = microscope clown!"},
        "risk": "critical", "risk_d": {"fr": "URGENCE", "ar": "طوارئ", "en": "EMERGENCY"},
        "advice": {"fr": "HOSPITALISATION ! ACT. Quinine IV si grave. Parasitemie /4-6h.",
                   "ar": "دخول المستشفى! ACT. كينين وريدي إذا خطير.",
                   "en": "HOSPITALIZATION! ACT. IV Quinine if severe. Parasitemia /4-6h."},
        "tests": ["TDR Paludisme", "Frottis+GE URGENCE", "Parasitemie quantitative", "NFS", "Bilan hepato-renal", "Glycemie"],
        "color": "#7f1d1d", "icon": "🚨",
        "cycle": {"fr": "Piqure anophele -> Sporozoites -> Hepatocytes -> Merozoites -> Hematies -> Gametocytes",
                  "ar": "لدغة الأنوفيل ثم سبوروزويت ثم خلايا كبدية ثم ميروزويت ثم كريات حمراء ثم خلايا جنسية",
                  "en": "Anopheles bite -> Sporozoites -> Hepatocytes -> Merozoites -> RBCs -> Gametocytes"},
        "keys": {"fr": "URGENCE moins de 2h\nFrottis: espece\nGE: 10x sensible\nPlus de 2 pourcent = grave\nBanane = P. falciparum",
                 "ar": "طوارئ أقل من ساعتين\nلطاخة: النوع\nGE: أكثر حساسية 10 مرات\nأكثر من 2 بالمئة = خطير\nموز = P. falciparum",
                 "en": "URGENT under 2h\nSmear: species\nThick drop: 10x sensitive\nOver 2 percent=severe\nBanana=P. falciparum"}
    },
    "Trypanosoma": {
        "sci": "Trypanosoma brucei gambiense / rhodesiense / cruzi",
        "morph": {"fr": "Forme S/C (15-30um), flagelle libre, membrane ondulante, kinetoplaste posterieur.",
                  "ar": "شكل S/C (15-30 ميكرومتر)، سوط حر، غشاء متموج، كينيتوبلاست خلفي.",
                  "en": "S/C shape (15-30um), free flagellum, undulating membrane, posterior kinetoplast."},
        "desc": {"fr": "Maladie du sommeil (tse-tse) ou Chagas (triatome). Phase hemolymphatique puis neurologique.",
                 "ar": "مرض النوم (تسي تسي) أو شاغاس (بق ثلاثي). مرحلة دموية ثم عصبية.",
                 "en": "Sleeping sickness (tsetse) or Chagas (triatomine). Hemolymphatic then neurological phase."},
        "funny": {"fr": "Il court avec sa membrane ondulante ! La tse-tse = le pire taxi !",
                  "ar": "يركض بغشائه المتموج! ذبابة تسي تسي = أسوأ تاكسي!",
                  "en": "Runs with its undulating membrane! Tsetse = worst taxi!"},
        "risk": "high", "risk_d": {"fr": "Eleve", "ar": "مرتفع", "en": "High"},
        "advice": {"fr": "Phase 1: Pentamidine. Phase 2: NECT/Melarsoprol. PL obligatoire.",
                   "ar": "المرحلة 1: بنتاميدين. المرحلة 2: NECT. بزل قطني إجباري.",
                   "en": "Phase 1: Pentamidine. Phase 2: NECT/Melarsoprol. LP mandatory."},
        "tests": ["Ponction lombaire", "Serologie CATT", "IgM", "Suc ganglionnaire", "NFS"],
        "color": "#ff0040", "icon": "🔴",
        "cycle": {"fr": "Piqure tse-tse -> Trypomastigotes -> Sang -> Phase 1 -> BHE -> Phase 2 neurologique",
                  "ar": "لدغة تسي تسي ثم تريبوماستيغوت ثم دم ثم مرحلة 1 ثم حاجز دماغي ثم مرحلة 2 عصبية",
                  "en": "Tsetse bite -> Trypomastigotes -> Blood -> Phase 1 -> BBB -> Phase 2 neurological"},
        "keys": {"fr": "Forme S/C + membrane ondulante\nKinetoplaste posterieur\nIgM tres elevee\nPL staging",
                 "ar": "شكل S/C + غشاء متموج\nكينيتوبلاست خلفي\nIgM مرتفع جدا\nتصنيف بالبزل القطني",
                 "en": "S/C+undulating membrane\nPosterior kinetoplast\nVery high IgM\nLP staging"}
    },
    "Schistosoma": {
        "sci": "Schistosoma haematobium / mansoni / japonicum",
        "morph": {"fr": "Oeuf ovoide (115-170um): eperon terminal (S. haematobium) ou lateral (S. mansoni). Miracidium mobile.",
                  "ar": "بيضة بيضاوية (115-170 ميكرومتر): شوكة طرفية (S. haematobium) أو جانبية (S. mansoni). ميراسيديوم متحرك.",
                  "en": "Ovoid egg (115-170um): terminal spine (S. haematobium) or lateral (S. mansoni). Motile miracidium."},
        "desc": {"fr": "Bilharziose. S. haematobium: uro-genitale. S. mansoni: hepato-intestinale.",
                 "ar": "البلهارسيا. S. haematobium: بولي تناسلي. S. mansoni: كبدي معوي.",
                 "en": "Schistosomiasis. S. haematobium: urogenital. S. mansoni: hepato-intestinal."},
        "funny": {"fr": "L'oeuf avec un dard ! Les cercaires = micro-torpilles !",
                  "ar": "البيضة ذات الشوكة! السركاريا = طوربيدات صغيرة!",
                  "en": "Egg with a stinger! Cercariae = micro-torpedoes!"},
        "risk": "medium", "risk_d": {"fr": "Moyen", "ar": "متوسط", "en": "Medium"},
        "advice": {"fr": "Praziquantel 40mg/kg dose unique. S. haematobium: urines de midi.",
                   "ar": "برازيكوانتيل 40 ملغ لكل كغ جرعة واحدة. S. haematobium: بول الظهيرة.",
                   "en": "Praziquantel 40mg/kg single dose. S. haematobium: midday urine."},
        "tests": ["ECBU midi", "Serologie", "Echo vesicale/hepatique", "NFS+eosinophilie", "Biopsie rectale"],
        "color": "#ff9500", "icon": "🟠",
        "cycle": {"fr": "Oeuf -> Miracidium -> Mollusque -> Cercaire -> Penetration cutanee -> Vers adultes -> Ponte",
                  "ar": "بيضة ثم ميراسيديوم ثم رخويات ثم سركاريا ثم اختراق الجلد ثم ديدان بالغة ثم وضع البيض",
                  "en": "Egg -> Miracidium -> Snail -> Cercaria -> Skin penetration -> Adult worms -> Egg laying"},
        "keys": {"fr": "S.h: eperon TERMINAL, urines MIDI\nS.m: eperon LATERAL, selles\nMiracidium vivant\nEosinophilie elevee",
                 "ar": "S.h: شوكة طرفية، بول الظهيرة\nS.m: شوكة جانبية، براز\nميراسيديوم حي\nفرط الحمضات",
                 "en": "S.h: TERMINAL spine, MIDDAY urine\nS.m: LATERAL spine, stool\nLiving miracidium\nHigh eosinophilia"}
    },
    "Negative": {
        "sci": "N/A",
        "morph": {"fr": "Absence d'elements parasitaires. Flore bacterienne normale.",
                  "ar": "غياب العناصر الطفيلية. فلورا بكتيرية طبيعية.",
                  "en": "No parasitic elements. Normal bacterial flora."},
        "desc": {"fr": "Echantillon negatif. Un seul negatif n'exclut pas (sensibilite 50-60%). Repeter x3.",
                 "ar": "عينة سلبية. فحص واحد سلبي لا يستبعد (حساسية 50-60%). كرر 3 مرات.",
                 "en": "Negative sample. Single negative doesn't exclude (50-60% sensitivity). Repeat x3."},
        "funny": {"fr": "Rien a signaler ! Mais les parasites sont des maitres du cache-cache !",
                  "ar": "لا شيء يذكر! لكن الطفيليات أساتذة في الاختباء!",
                  "en": "Nothing to report! But parasites are hide-and-seek masters!"},
        "risk": "none", "risk_d": {"fr": "Negatif", "ar": "سلبي", "en": "Negative"},
        "advice": {"fr": "RAS. Repeter x3 si suspicion clinique.",
                   "ar": "لا شيء. كرر 3 مرات إذا كان هناك اشتباه.",
                   "en": "Clear. Repeat x3 if clinical suspicion."},
        "tests": ["Repeter EPS x3", "Serologie ciblee", "NFS (eosinophilie?)"],
        "color": "#00ff88", "icon": "🟢",
        "cycle": {"fr": "N/A", "ar": "غير متوفر", "en": "N/A"},
        "keys": {"fr": "Direct+Lugol negatif\nConcentration negative\nRepeter x3",
                 "ar": "مباشر+لوغول سلبي\nتركيز سلبي\nكرر 3 مرات",
                 "en": "Direct+Lugol negative\nConcentration negative\nRepeat x3"}
    }
}

CLASS_NAMES = list(PARASITE_DB.keys())


# ============================================
#  QUIZ QUESTIONS (60+)
# ============================================
def mq(fr, ar, en):
    return {"fr": fr, "ar": ar, "en": en}


QUIZ_QUESTIONS = [
    {"q": mq("Quel parasite presente une bague a chaton dans les hematies?", "أي طفيلي يظهر شكل الخاتم في كريات الدم الحمراء؟", "Which parasite shows a signet ring in RBCs?"),
     "opts": ["Giardia", "Plasmodium", "Leishmania", "Amoeba"], "ans": 1, "cat": "Hematozoaires",
     "expl": mq("Plasmodium: bague a chaton au stade trophozoite jeune.", "البلازموديوم: شكل الخاتم في مرحلة الطور النشط.", "Plasmodium: signet ring at young trophozoite stage.")},
    {"q": mq("Le kyste mature de Giardia possede combien de noyaux?", "كم عدد أنوية كيس الجيارديا الناضج؟", "How many nuclei does a mature Giardia cyst have?"),
     "opts": ["2", "4", "6", "8"], "ans": 1, "cat": "Protozoaires",
     "expl": mq("4 noyaux. Le trophozoite en a 2.", "4 أنوية. الطور النشط له نواتان.", "4 nuclei. Trophozoite has 2.")},
    {"q": mq("Quel parasite est transmis par le phlebotome?", "أي طفيلي ينتقل عبر ذبابة الرمل؟", "Which parasite is sandfly-transmitted?"),
     "opts": ["Plasmodium", "Trypanosoma", "Leishmania", "Schistosoma"], "ans": 2, "cat": "Vecteurs",
     "expl": mq("Leishmania = phlebotome.", "ليشمانيا = ذبابة الرمل.", "Leishmania = sandfly.")},
    {"q": mq("L'eperon terminal caracterise:", "الشوكة الطرفية تميز:", "Terminal spine characterizes:"),
     "opts": ["Ascaris", "S. haematobium", "S. mansoni", "Taenia"], "ans": 1, "cat": "Helminthes",
     "expl": mq("S. haematobium=terminal, S. mansoni=lateral.", "S. haematobium=طرفية, S. mansoni=جانبية.", "S. haematobium=terminal, S. mansoni=lateral.")},
    {"q": mq("Examen urgent suspicion paludisme?", "الفحص الطارئ عند الاشتباه بالملاريا؟", "Urgent exam for malaria?"),
     "opts": ["Coproculture", "ECBU", "Goutte epaisse+Frottis", "Serologie"], "ans": 2, "cat": "Diagnostic",
     "expl": mq("GE+Frottis = reference urgente.", "قطرة سميكة+لطاخة = المرجع.", "Thick drop+Smear = urgent reference.")},
    {"q": mq("E. histolytica se distingue par:", "يتميز E. histolytica بـ:", "E. histolytica distinguished by:"),
     "opts": ["Flagelles", "Hematies phagocytees", "Membrane ondulante", "Kinetoplaste"], "ans": 1, "cat": "Morphologie",
     "expl": mq("Hematies phagocytees = pathogenicite.", "الكريات المبتلعة = معيار المرضية.", "Phagocytosed RBCs = pathogenicity.")},
    {"q": mq("Chagas est causee par:", "مرض شاغاس يسببه:", "Chagas is caused by:"),
     "opts": ["T. b. gambiense", "T. cruzi", "L. donovani", "P. vivax"], "ans": 1, "cat": "Protozoaires",
     "expl": mq("T. cruzi transmis par triatomes.", "T. cruzi عبر البق الثلاثي.", "T. cruzi by triatomines.")},
    {"q": mq("Colorant pour amastigotes Leishmania?", "ملون أماستيغوت الليشمانيا؟", "Stain for Leishmania amastigotes?"),
     "opts": ["Ziehl-Neelsen", "Gram", "MGG/Giemsa", "Lugol"], "ans": 2, "cat": "Techniques",
     "expl": mq("MGG = noyau + kinetoplaste.", "MGG = نواة + كينيتوبلاست.", "MGG = nucleus + kinetoplast.")},
    {"q": mq("Traitement reference bilharziose?", "العلاج المرجعي للبلهارسيا؟", "Reference treatment schistosomiasis?"),
     "opts": ["Chloroquine", "Metronidazole", "Praziquantel", "Albendazole"], "ans": 2, "cat": "Therapeutique",
     "expl": mq("Praziquantel = choix numero 1.", "برازيكوانتيل = الخيار الأول.", "Praziquantel = 1st choice.")},
    {"q": mq("Face de hibou observee chez:", "وجه البومة عند:", "Owl face observed in:"),
     "opts": ["Plasmodium", "Giardia", "Amoeba", "Trypanosoma"], "ans": 1, "cat": "Morphologie",
     "expl": mq("2 noyaux symetriques Giardia.", "نواتان متماثلتان للجيارديا.", "2 symmetrical Giardia nuclei.")},
    {"q": mq("Technique de Ritchie:", "تقنية ريتشي:", "Ritchie technique:"),
     "opts": ["Coloration", "Concentration diphasique", "Culture", "Serologie"], "ans": 1, "cat": "Techniques",
     "expl": mq("Formol-ether = concentration.", "فورمول-إيثر = تركيز.", "Formalin-ether = concentration.")},
    {"q": mq("Le Lugol met en evidence:", "اللوغول يظهر:", "Lugol highlights:"),
     "opts": ["Flagelles", "Noyaux des kystes", "Hematies", "Bacteries"], "ans": 1, "cat": "Techniques",
     "expl": mq("Iode colore glycogene et noyaux.", "اليود يلون الغليكوجين والأنوية.", "Iodine stains glycogen+nuclei.")},
    {"q": mq("x100 necessite:", "العدسة x100 تحتاج:", "x100 requires:"),
     "opts": ["Eau", "Huile d'immersion", "Alcool", "Serum"], "ans": 1, "cat": "Microscopie",
     "expl": mq("Huile augmente indice refraction.", "الزيت يزيد معامل الانكسار.", "Oil increases refractive index.")},
    {"q": mq("Scotch-test Graham recherche:", "اختبار غراهام يبحث عن:", "Graham test looks for:"),
     "opts": ["Giardia", "Enterobius", "Ascaris", "Taenia"], "ans": 1, "cat": "Techniques",
     "expl": mq("Oeufs d'oxyure peri-anaux.", "بيض الأكسيور حول الشرج.", "Pinworm eggs perianal.")},
    {"q": mq("Coloration Cryptosporidium?", "تلوين الكريبتوسبوريديوم؟", "Cryptosporidium staining?"),
     "opts": ["Lugol", "Ziehl-Neelsen modifie", "MGG", "Gram"], "ans": 1, "cat": "Techniques",
     "expl": mq("ZN modifie = oocystes roses.", "ZN معدل = أكياس بيضية وردية.", "Modified ZN = pink oocysts.")},
    {"q": mq("Oeuf d'Ascaris:", "بيضة الأسكاريس:", "Ascaris egg:"),
     "opts": ["Avec eperon", "Mamelonne", "Opercule", "En citron"], "ans": 1, "cat": "Helminthes",
     "expl": mq("Ovoide, mamelonne, coque epaisse.", "بيضاوي، حليمي، قشرة سميكة.", "Ovoid, mammillated, thick shell.")},
    {"q": mq("Scolex T. solium possede:", "رأس T. solium يحتوي:", "T. solium scolex has:"),
     "opts": ["Ventouses seules", "Crochets seuls", "Ventouses+crochets", "Bothridies"], "ans": 2, "cat": "Helminthes",
     "expl": mq("Arme = ventouses + crochets.", "مسلحة = ممصات + خطاطيف.", "Armed = suckers + hooks.")},
    {"q": mq("Eosinophilie sanguine oriente vers:", "فرط الحمضات يوجه نحو:", "Eosinophilia points to:"),
     "opts": ["Bacteries", "Helminthiase", "Virose", "Paludisme"], "ans": 1, "cat": "Diagnostic",
     "expl": mq("Eosinophilie = helminthiase.", "فرط الحمضات = ديدان.", "Eosinophilia = helminthiasis.")},
    {"q": mq("Vecteur du paludisme?", "ناقل الملاريا؟", "Malaria vector?"),
     "opts": ["Aedes", "Culex", "Anopheles", "Simulium"], "ans": 2, "cat": "Epidemiologie",
     "expl": mq("Anophele femelle.", "أنثى الأنوفيل.", "Female Anopheles.")},
    {"q": mq("Kyste hydatique du a:", "الكيس العداري يسببه:", "Hydatid cyst caused by:"),
     "opts": ["T. saginata", "E. granulosus", "Fasciola", "Toxocara"], "ans": 1, "cat": "Helminthes",
     "expl": mq("Echinococcus granulosus (chien).", "Echinococcus granulosus (كلب).", "Echinococcus granulosus (dog).")},
    {"q": mq("Membrane ondulante:", "الغشاء المتموج:", "Undulating membrane:"),
     "opts": ["Giardia", "Trypanosoma", "Leishmania", "Plasmodium"], "ans": 1, "cat": "Morphologie",
     "expl": mq("Trypanosoma = membrane ondulante.", "تريبانوسوما = غشاء متموج.", "Trypanosoma = undulating membrane.")},
    {"q": mq("Gametocyte banane:", "خلية جنسية موز:", "Banana gametocyte:"),
     "opts": ["P. vivax", "P. falciparum", "P. malariae", "P. ovale"], "ans": 1, "cat": "Hematozoaires",
     "expl": mq("Pathognomonique P. falciparum.", "مميز لـ P. falciparum.", "Pathognomonic P. falciparum.")},
    {"q": mq("Kyste E. coli: noyaux?", "كيس E. coli: أنوية؟", "E. coli cyst: nuclei?"),
     "opts": ["4", "6", "8", "12"], "ans": 2, "cat": "Morphologie",
     "expl": mq("E. coli=8, E. histolytica=4.", "E. coli=8, E. histolytica=4.", "E. coli=8, E. histolytica=4.")},
    {"q": mq("Metronidazole inefficace contre:", "ميترونيدازول غير فعال ضد:", "Metronidazole ineffective against:"),
     "opts": ["E. histolytica", "Giardia", "Helminthes", "Trichomonas"], "ans": 2, "cat": "Therapeutique",
     "expl": mq("Anti-protozoaire, pas anti-helminthique.", "مضاد أوليات، ليس مضاد ديدان.", "Anti-protozoal, not anti-helminthic.")},
    {"q": mq("Albendazole est:", "الألبندازول:", "Albendazole is:"),
     "opts": ["Anti-protozoaire", "Anti-helminthique", "Antibiotique", "Antifongique"], "ans": 1, "cat": "Therapeutique",
     "expl": mq("Large spectre helminthes.", "واسع الطيف ضد الديدان.", "Broad spectrum anti-helminthic.")},
    {"q": mq("Paludisme grave:", "ملاريا خطيرة:", "Severe malaria:"),
     "opts": ["Chloroquine", "Artesunate IV", "Metronidazole", "Praziquantel"], "ans": 1, "cat": "Therapeutique",
     "expl": mq("Artesunate IV = 1ere ligne OMS.", "أرتيسونات وريدي = الخط الأول.", "IV Artesunate = WHO 1st line.")},
    {"q": mq("Fievre+frissons retour Afrique?", "حمى+قشعريرة بعد العودة من إفريقيا؟", "Fever+chills returning from Africa?"),
     "opts": ["Amibiase", "Paludisme", "Bilharziose", "Giardiose"], "ans": 1, "cat": "Cas clinique",
     "expl": mq("Paludisme jusqu'a preuve du contraire.", "ملاريا حتى يثبت العكس.", "Malaria until proven otherwise.")},
    {"q": mq("Hematurie+baignade eau douce:", "بيلة دموية+سباحة ماء عذب:", "Hematuria+freshwater swimming:"),
     "opts": ["Giardiose", "Paludisme", "Bilharziose urinaire", "Amibiase"], "ans": 2, "cat": "Cas clinique",
     "expl": mq("S. haematobium.", "S. haematobium.", "S. haematobium.")},
    {"q": mq("Diarrhee graisseuse enfant:", "إسهال دهني عند طفل:", "Greasy diarrhea child:"),
     "opts": ["Amibiase", "Giardiose", "Cryptosporidiose", "Salmonellose"], "ans": 1, "cat": "Cas clinique",
     "expl": mq("Giardia = malabsorption frequente.", "الجيارديا = سوء امتصاص شائع.", "Giardia = common malabsorption.")},
    {"q": mq("Toxoplasma: hote definitif?", "التوكسوبلازما: المضيف النهائي؟", "Toxoplasma: definitive host?"),
     "opts": ["Homme", "Chat", "Chien", "Moustique"], "ans": 1, "cat": "Epidemiologie",
     "expl": mq("Chat = cycle sexue.", "القط = الدورة الجنسية.", "Cat = sexual cycle.")},
    {"q": mq("Willis utilise:", "تقنية ويليس تستخدم:", "Willis uses:"),
     "opts": ["Formol-ether", "NaCl sature (flottation)", "Acide-alcool", "Lugol"], "ans": 1, "cat": "Techniques",
     "expl": mq("Flottation NaCl sature.", "تعويم في NaCl مشبع.", "Saturated NaCl flotation.")},
]


# ============================================
#  CHATBOT KNOWLEDGE BASE
# ============================================
CHAT_KB = {}
for pname, pdata in PARASITE_DB.items():
    if pname == "Negative":
        continue
    key = pname.lower().split("(")[0].strip().split(" ")[0].lower()
    CHAT_KB[key] = {
        "fr": f"**{pname}** ({pdata['sci']})\n\n**Morphologie:** {pdata['morph'].get('fr','')}\n\n**Description:** {pdata['desc'].get('fr','')}\n\n**Traitement:** {pdata['advice'].get('fr','')}\n\n**Examens:** {', '.join(pdata.get('tests',[]))}\n\n{pdata['funny'].get('fr','')}",
        "ar": f"**{pname}** ({pdata['sci']})\n\n**المورفولوجيا:** {pdata['morph'].get('ar','')}\n\n**الوصف:** {pdata['desc'].get('ar','')}\n\n**العلاج:** {pdata['advice'].get('ar','')}\n\n{pdata['funny'].get('ar','')}",
        "en": f"**{pname}** ({pdata['sci']})\n\n**Morphology:** {pdata['morph'].get('en','')}\n\n**Description:** {pdata['desc'].get('en','')}\n\n**Treatment:** {pdata['advice'].get('en','')}\n\n{pdata['funny'].get('en','')}"
    }
    sci_key = pdata["sci"].lower().split("/")[0].strip().split(" ")[-1]
    if sci_key not in CHAT_KB:
        CHAT_KB[sci_key] = CHAT_KB[key]

CHAT_KB.update({
    "amibe": CHAT_KB.get("amoeba", {}),
    "malaria": CHAT_KB.get("plasmodium", {}),
    "paludisme": CHAT_KB.get("plasmodium", {}),
    "bilharziose": CHAT_KB.get("schistosoma", {}),
    "microscope": {
        "fr": "**Microscopie:**\n\n- **x10:** Reperage\n- **x40:** Oeufs/kystes\n- **x100 (immersion):** Plasmodium, Leishmania\n\nNettoyer l'objectif x100 apres l'huile!\n\n**Types:** Optique, Fluorescence, Contraste de phase, Fond noir, Confocal",
        "ar": "**المجهرية:**\n\n- **x10:** استطلاع\n- **x40:** بيض/أكياس\n- **x100 (غمر):** بلازموديوم، ليشمانيا\n\nنظف العدسة x100 بعد الزيت!",
        "en": "**Microscopy:**\n\n- **x10:** Survey\n- **x40:** Eggs/cysts\n- **x100 (immersion):** Plasmodium, Leishmania\n\nClean x100 after oil!"
    },
    "coloration": {
        "fr": "**Colorations:**\n\n- **Lugol:** Noyaux kystes, glycogene\n- **MGG/Giemsa:** Parasites sanguins\n- **Ziehl-Neelsen modifie:** Cryptosporidium\n- **Trichrome:** Parasites intestinaux\n- **Hematoxyline ferrique:** Amibes\n\nLugol frais chaque semaine!",
        "ar": "**التلوينات:**\n\n- **لوغول:** أنوية الأكياس\n- **MGG/جيمزا:** طفيليات الدم\n- **ZN معدل:** كريبتوسبوريديوم",
        "en": "**Staining:**\n\n- **Lugol:** Cyst nuclei\n- **MGG/Giemsa:** Blood parasites\n- **Modified ZN:** Cryptosporidium"
    },
    "concentration": {
        "fr": "**Techniques concentration:**\n\n- **Ritchie (Formol-ether):** REFERENCE\n- **Willis (NaCl sature):** Flottation\n- **Kato-Katz:** Semi-quantitatif\n- **Baermann:** Larves Strongyloides\n- **MIF:** Fixation+coloration",
        "ar": "**تقنيات التركيز:**\n\n- **ريتشي:** المرجع\n- **ويليس:** تعويم\n- **كاتو-كاتز:** شبه كمي",
        "en": "**Concentration:**\n\n- **Ritchie:** REFERENCE\n- **Willis:** Flotation\n- **Kato-Katz:** Semi-quantitative"
    },
    "selle": {
        "fr": "**EPS Complet:**\n\n1. **Macroscopique:** Consistance, couleur, sang, mucus\n2. **Direct:** NaCl 0.9% + Lugol\n3. **Concentration:** Ritchie/Willis\n\nExaminer dans 30 min!\nRepeter x3!\n\nSelles liquides = trophozoites, Formees = kystes",
        "ar": "**فحص البراز:**\n\n1. عياني\n2. مباشر: NaCl + لوغول\n3. تركيز: ريتشي/ويليس\n\nفحص خلال 30 دقيقة!\nكرر 3 مرات!",
        "en": "**Complete Stool Exam:**\n\n1. Macroscopic\n2. Direct: NaCl + Lugol\n3. Concentration: Ritchie/Willis\n\nExamine within 30 min!\nRepeat x3!"
    },
    "hygiene": {
        "fr": "**Prevention:**\n\n- Lavage mains 30s\n- Eau potable\n- Cuisson viande plus de 65C\n- Moustiquaires\n- Eviter eaux stagnantes\n- Lavage fruits/legumes\n\n80% des parasitoses sont evitables!",
        "ar": "**الوقاية:**\n\n- غسل اليدين 30 ثانية\n- ماء صالح للشرب\n- طهي اللحم أكثر من 65 درجة\n- ناموسيات\n\n80% من الطفيليات يمكن الوقاية منها!",
        "en": "**Prevention:**\n\n- Handwashing 30s\n- Safe water\n- Cook meat over 65C\n- Mosquito nets\n\n80% of parasitoses are preventable!"
    },
    "traitement": {
        "fr": "**Traitements:**\n\n- **Metronidazole:** Amoeba+Giardia+Trichomonas\n- **Albendazole:** Helminthes large spectre\n- **Praziquantel:** Schistosoma+Cestodes\n- **Artesunate/ACT:** Paludisme\n- **Glucantime:** Leishmaniose cutanee\n- **Ivermectine:** Filarioses\n- **Niclosamide:** Tenias",
        "ar": "**العلاجات:**\n\n- **ميترونيدازول:** أميبا+جيارديا\n- **ألبندازول:** ديدان\n- **برازيكوانتيل:** بلهارسيا+شريطيات\n- **أرتيسونات:** ملاريا",
        "en": "**Treatments:**\n\n- **Metronidazole:** Amoeba+Giardia+Trichomonas\n- **Albendazole:** Broad spectrum helminths\n- **Praziquantel:** Schistosoma+Cestodes\n- **Artesunate/ACT:** Malaria"
    },
    "toxoplasma": {
        "fr": "**Toxoplasma gondii**\n\n- Tachyzoite en arc (4-8um)\n- Hote definitif: **Chat**\n- DANGER femme enceinte seronegative!\n- Diagnostic: Serologie IgG/IgM, avidite\n- Prevention: Cuisson viande, lavage crudites, eviter litiere",
        "ar": "**توكسوبلازما**\n\nالمضيف النهائي: القط\nخطر على الحامل!\nالتشخيص: مصلية IgG/IgM",
        "en": "**Toxoplasma gondii**\n\n- Definitive host: **Cat**\n- DANGER seronegative pregnant!\n- Diagnosis: IgG/IgM serology"
    },
    "ascaris": {
        "fr": "**Ascaris lumbricoides**\n\n- Ver rond 15-35cm!\n- Oeuf mamelonne 60-70um\n- Cycle: Migration hepato-pulmonaire\n- Syndrome Loffler: Toux+eosinophilie\n- Traitement: Albendazole 400mg",
        "ar": "**الأسكاريس**\n\nدودة 15-35 سم!\nبيضة حليمية\nالعلاج: ألبندازول 400 ملغ",
        "en": "**Ascaris lumbricoides**\n\n- Round worm 15-35cm!\n- Mammillated egg 60-70um\n- Treatment: Albendazole 400mg"
    },
    "taenia": {
        "fr": "**Taenia**\n\n- **T. saginata:** Boeuf, inerme\n- **T. solium:** Porc, arme -> cysticercose!\n- Diagnostic: Anneaux dans selles\n- Traitement: Praziquantel/Niclosamide",
        "ar": "**الشريطية**\n\nT. saginata: بقر\nT. solium: خنزير -> كيسات مذنبة!\nالعلاج: برازيكوانتيل",
        "en": "**Taenia**\n\n- T. saginata: Beef\n- T. solium: Pork -> cysticercosis!\n- Treatment: Praziquantel"
    },
    "oxyure": {
        "fr": "**Oxyure (Enterobius)**\n\n- Ver blanc environ 1cm\n- Prurit anal nocturne enfant\n- Scotch-test MATIN avant toilette!\n- Traitement: Flubendazole + traiter TOUTE la famille!",
        "ar": "**الأكسيور**\n\nدودة بيضاء حوالي 1 سم\nحكة شرجية ليلية\nاختبار غراهام صباحا!\nعلاج كل العائلة!",
        "en": "**Pinworm (Enterobius)**\n\n- White worm about 1cm\n- Nocturnal anal pruritus in children\n- Graham test MORNING!\n- Treat WHOLE family!"
    },
    "cryptosporidium": {
        "fr": "**Cryptosporidium**\n\n- Oocyste 4-6um (tres petit!)\n- ZN modifie = rose sur vert\n- Diarrhee immunodeprime (VIH)\n- Traitement: Nitazoxanide",
        "ar": "**كريبتوسبوريديوم**\n\nكيس 4-6 ميكرومتر\nZN معدل\nإسهال عند ناقصي المناعة",
        "en": "**Cryptosporidium**\n\n- Oocyst 4-6um\n- Modified ZN\n- Diarrhea in immunocompromised\n- Nitazoxanide"
    },
    "bonjour": {"fr": "Bonjour! Comment puis-je vous aider?", "ar": "مرحبا! كيف أقدر أساعدك؟", "en": "Hello! How can I help?"},
    "salut": {"fr": "Salut! Posez votre question!", "ar": "مرحبا!", "en": "Hi!"},
    "hello": {"fr": "Hello!", "ar": "مرحبا!", "en": "Hello! Type any parasite name!"},
    "merci": {"fr": "De rien!", "ar": "عفوا!", "en": "You're welcome!"},
    "aide": {
        "fr": "**Je connais:**\n\nParasites: Amoeba, Giardia, Leishmania, Plasmodium, Trypanosoma, Schistosoma, Toxoplasma, Ascaris, Taenia, Oxyure, Cryptosporidium...\n\nTechniques: microscope, coloration, concentration, selle\n\nTraitements: traitement\n\nPrevention: hygiene\n\nTapez un mot-cle!",
        "ar": "**أعرف:**\n\nالطفيليات: الأميبا، الجيارديا، البلازموديوم، الليشمانيا...\n\nالتقنيات: microscope, coloration, concentration\n\nاكتب كلمة مفتاحية!",
        "en": "**I know:**\n\nParasites: Amoeba, Giardia, Leishmania, Plasmodium, Trypanosoma, Schistosoma, Toxoplasma, Ascaris...\n\nTechniques: microscope, staining, concentration\n\nType a keyword!"
    },
    "help": {"fr": "Tapez aide!", "ar": "اكتب مساعدة!", "en": "Type any parasite name or technique!"},
})


def chatbot_reply(msg):
    lang = st.session_state.get("lang", "fr")
    low = msg.lower().strip()
    for key, resp in CHAT_KB.items():
        if key in low:
            if isinstance(resp, dict):
                return resp.get(lang, resp.get("fr", str(resp)))
            return str(resp)
    for pname, pdata in PARASITE_DB.items():
        if pname == "Negative":
            continue
        checks = [pname.lower(), pdata["sci"].lower()]
        for check in checks:
            for word in check.split():
                if len(word) > 3 and word in low:
                    entry = CHAT_KB.get(pname.lower().split("(")[0].strip().split(" ")[0].lower())
                    if entry:
                        return entry.get(lang, entry.get("fr", ""))
    return t("chat_not_found")


# ============================================
#  DAILY TIPS
# ============================================
TIPS = {
    "fr": [
        "Examiner les selles dans les 30 min pour voir les trophozoites mobiles.",
        "Le Lugol met en evidence les noyaux des kystes. Fraichement prepare!",
        "Frottis mince: angle 45 degres pour monocouche parfaite.",
        "Goutte epaisse = 10x plus sensible que frottis mince.",
        "Urines de midi pour S. haematobium (pic d'excretion).",
        "Repeter EPS x3 a quelques jours d'intervalle.",
        "Metronidazole = Amoeba + Giardia + Trichomonas !",
        "ZN modifie indispensable pour Cryptosporidium.",
        "1ere GE negative ne suffit pas. Repeter a 6-12h.",
        "Eosinophilie = helminthiase. Toujours verifier!",
    ],
    "ar": [
        "افحص البراز خلال 30 دقيقة لرؤية الأطوار النشطة المتحركة.",
        "اللوغول يظهر أنوية الأكياس. تحضير طازج!",
        "القطرة السميكة أكثر حساسية 10 مرات من اللطاخة.",
        "بول الظهيرة لـ S. haematobium.",
        "كرر EPS 3 مرات بفاصل عدة أيام.",
        "ميترونيدازول = أميبا + جيارديا + تريكوموناس!",
    ],
    "en": [
        "Examine stool within 30 min to see motile trophozoites.",
        "Lugol highlights cyst nuclei. Freshly prepared!",
        "Thick drop = 10x more sensitive than thin smear.",
        "Midday urine for S. haematobium.",
        "Repeat stool exam x3 at intervals.",
        "Metronidazole = Amoeba + Giardia + Trichomonas!",
    ]
}

# ============================================
#  SESSION STATE
# ============================================
DEFAULTS = {
    "logged_in": False, "user_id": None, "user_name": "", "user_role": "viewer",
    "user_full_name": "", "lang": "fr",
    "demo_seed": None, "heatmap_seed": None,
    "quiz_state": {"current": 0, "score": 0, "answered": [], "active": False, "order": [], "wrong": []},
    "chat_history": [],
    "voice_text": None, "voice_lang": None,
}
for k, v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ============================================
#  UTILITY FUNCTIONS
# ============================================
def has_role(lvl):
    return ROLES.get(st.session_state.get("user_role", "viewer"), {}).get("level", 0) >= lvl


def risk_color(lv):
    return {"critical": "#ff0040", "high": "#ff3366", "medium": "#ff9500", "low": "#00e676", "none": "#00ff88"}.get(lv, "#888")


def risk_pct(lv):
    return {"critical": 100, "high": 80, "medium": 50, "low": 25, "none": 0}.get(lv, 0)


# ============================================
#  VOICE SYSTEM (Web Speech API - Fixed)
# ============================================
def render_voice_player():
    """Renders hidden audio player using Web Speech API"""
    if st.session_state.get("voice_text"):
        text = st.session_state.voice_text.replace("'", "\\'").replace('"', '\\"').replace('\n', ' ').replace('\r', ' ')
        lang_code = {"fr": "fr-FR", "ar": "ar-SA", "en": "en-US"}.get(
            st.session_state.get("voice_lang", st.session_state.get("lang", "fr")), "fr-FR")
        
        html_code = f"""
        <div id="voice-container" style="display:none;">
            <script>
            (function() {{
                try {{
                    if ('speechSynthesis' in window) {{
                        window.speechSynthesis.cancel();
                        setTimeout(function() {{
                            var utterance = new SpeechSynthesisUtterance('{text}');
                            utterance.lang = '{lang_code}';
                            utterance.rate = 0.9;
                            utterance.pitch = 1.0;
                            utterance.volume = 1.0;
                            
                            var voices = window.speechSynthesis.getVoices();
                            if (voices.length > 0) {{
                                for (var i = 0; i < voices.length; i++) {{
                                    if (voices[i].lang.startsWith('{lang_code.split("-")[0]}')) {{
                                        utterance.voice = voices[i];
                                        break;
                                    }}
                                }}
                            }}
                            
                            window.speechSynthesis.speak(utterance);
                        }}, 200);
                    }}
                }} catch(e) {{
                    console.log('Speech error:', e);
                }}
            }})();
            </script>
        </div>
        """
        st.components.v1.html(html_code, height=0)
        st.session_state.voice_text = None
        st.session_state.voice_lang = None


def speak(text, lang=None):
    """Queue text for speaking"""
    st.session_state.voice_text = text
    st.session_state.voice_lang = lang or st.session_state.get("lang", "fr")


def stop_speech():
    """Stop speech"""
    st.session_state.voice_text = None
    st.components.v1.html("""
    <script>
    try { window.speechSynthesis.cancel(); } catch(e) {}
    </script>
    """, height=0)


# ============================================
#  AI ENGINE
# ============================================
@st.cache_resource(show_spinner=False)
def load_model():
    m, mn, mt = None, None, None
    try:
        import tensorflow as tf
        for ext in [".keras", ".h5"]:
            ff = [f for f in os.listdir(".") if f.endswith(ext) and os.path.isfile(f)]
            if ff:
                mn = ff[0]
                m = tf.keras.models.load_model(mn, compile=False)
                mt = "keras"
                break
        if m is None:
            ff = [f for f in os.listdir(".") if f.endswith(".tflite") and os.path.isfile(f)]
            if ff:
                mn = ff[0]
                m = tf.lite.Interpreter(model_path=mn)
                m.allocate_tensors()
                mt = "tflite"
    except Exception:
        pass
    return m, mn, mt


def predict(model, mt, img, seed=None):
    res = {"label": "Negative", "conf": 0, "preds": {}, "rel": False, "demo": False, "risk": "none"}
    rm = {"Plasmodium": "critical", "Amoeba (E. histolytica)": "high", "Leishmania": "high",
          "Trypanosoma": "high", "Giardia": "medium", "Schistosoma": "medium", "Negative": "none"}
    if model is None:
        res["demo"] = True
        if seed is None:
            seed = random.randint(0, 999999)
        rng = random.Random(seed)
        lb = rng.choice(CLASS_NAMES)
        cf = rng.randint(55, 98)
        ap = {}
        rem = 100.0 - cf
        for c in CLASS_NAMES:
            ap[c] = float(cf) if c == lb else round(rng.uniform(0, rem / max(1, len(CLASS_NAMES) - 1)), 1)
        res.update({"label": lb, "conf": cf, "preds": ap, "rel": cf >= CONFIDENCE_THRESHOLD, "risk": rm.get(lb, "none")})
        return res
    try:
        import tensorflow as tf
        im = ImageOps.fit(img.convert("RGB"), MODEL_INPUT_SIZE, Image.LANCZOS)
        arr = np.expand_dims(np.asarray(im).astype(np.float32) / 127.5 - 1.0, 0)
        if mt == "tflite":
            inp, out = model.get_input_details(), model.get_output_details()
            model.set_tensor(inp[0]['index'], arr)
            model.invoke()
            pr = model.get_tensor(out[0]['index'])[0]
        else:
            pr = model.predict(arr, verbose=0)[0]
        ix = int(np.argmax(pr))
        cf = int(pr[ix] * 100)
        lb = CLASS_NAMES[ix] if ix < len(CLASS_NAMES) else "Negative"
        ap = {CLASS_NAMES[i]: round(float(pr[i]) * 100, 1) for i in range(min(len(pr), len(CLASS_NAMES)))}
        res.update({"label": lb, "conf": cf, "preds": ap, "rel": cf >= CONFIDENCE_THRESHOLD, "risk": rm.get(lb, "none")})
    except Exception as e:
        st.error(f"Error: {e}")
    return res


# ============================================
#  IMAGE PROCESSING
# ============================================
def gen_heatmap(img, seed=None):
    im = img.copy().convert("RGB")
    w, h = im.size
    if seed is None:
        seed = hash(im.tobytes()[:1000]) % 1000000
    rng = random.Random(seed)
    ea = np.array(ImageOps.grayscale(im).filter(ImageFilter.FIND_EDGES))
    bs, mx, bx, by = 32, 0, w // 2, h // 2
    for y in range(0, h - bs, bs // 2):
        for x in range(0, w - bs, bs // 2):
            s = np.mean(ea[y:y + bs, x:x + bs])
            if s > mx:
                mx, bx, by = s, x + bs // 2, y + bs // 2
    bx = max(50, min(w - 50, bx + rng.randint(-w // 10, w // 10)))
    by = max(50, min(h - 50, by + rng.randint(-h // 10, h // 10)))
    hm = Image.new('RGBA', (w, h), (0, 0, 0, 0))
    d = ImageDraw.Draw(hm)
    mr = min(w, h) // 3
    for r in range(mr, 0, -2):
        a = max(0, min(200, int(200 * (1 - r / mr))))
        rat = r / mr
        if rat > 0.65:
            c = (0, 255, 100, a // 4)
        elif rat > 0.35:
            c = (255, 255, 0, a // 2)
        else:
            c = (255, 0, 60, a)
        d.ellipse([bx - r, by - r, bx + r, by + r], fill=c)
    return Image.alpha_composite(im.convert('RGBA'), hm).convert('RGB')


def thermal(img):
    return ImageOps.colorize(
        ImageOps.grayscale(ImageEnhance.Contrast(img).enhance(1.5)).filter(ImageFilter.GaussianBlur(1)),
        black="navy", white="yellow", mid="red")


def edges_filter(img):
    return ImageOps.grayscale(img).filter(ImageFilter.FIND_EDGES)


def enhanced_filter(img):
    return ImageEnhance.Contrast(ImageEnhance.Sharpness(img).enhance(2.0)).enhance(2.0)


def negative_filter(img):
    return ImageOps.invert(img.convert("RGB"))


def emboss_filter(img):
    return img.filter(ImageFilter.EMBOSS)


def adjust(img, br=1.0, co=1.0, sa=1.0):
    r = img.copy()
    if br != 1.0:
        r = ImageEnhance.Brightness(r).enhance(br)
    if co != 1.0:
        r = ImageEnhance.Contrast(r).enhance(co)
    if sa != 1.0:
        r = ImageEnhance.Color(r).enhance(sa)
    return r


def zoom_img(img, lv):
    if lv <= 1.0:
        return img
    w, h = img.size
    nw, nh = int(w / lv), int(h / lv)
    l, tp = (w - nw) // 2, (h - nh) // 2
    return img.crop((l, tp, l + nw, tp + nh)).resize((w, h), Image.LANCZOS)


def compare_imgs(i1, i2):
    a1 = np.array(i1.convert("RGB").resize((128, 128))).astype(float)
    a2 = np.array(i2.convert("RGB").resize((128, 128))).astype(float)
    mse = np.mean((a1 - a2) ** 2)
    m1, m2, s1, s2 = np.mean(a1), np.mean(a2), np.std(a1), np.std(a2)
    s12 = np.mean((a1 - m1) * (a2 - m2))
    c1, c2 = (0.01 * 255) ** 2, (0.03 * 255) ** 2
    ssim = ((2 * m1 * m2 + c1) * (2 * s12 + c2)) / ((m1 ** 2 + m2 ** 2 + c1) * (s1 ** 2 + s2 ** 2 + c2))
    return {"mse": round(mse, 2), "ssim": round(float(ssim), 4), "sim": round(float(ssim) * 100, 1)}


def pixel_diff(i1, i2):
    a1 = np.array(i1.convert("RGB").resize((256, 256))).astype(float)
    a2 = np.array(i2.convert("RGB").resize((256, 256))).astype(float)
    diff = np.abs(a1 - a2).astype(np.uint8)
    diff = np.clip(diff * 3, 0, 255).astype(np.uint8)
    return Image.fromarray(diff)


def histogram(img):
    r, g, b = img.convert("RGB").split()
    return {"red": list(r.histogram()), "green": list(g.histogram()), "blue": list(b.histogram())}


# ============================================
#  PDF - Enhanced Professional
# ============================================
def _sp(text):
    if not text:
        return ""
    reps = {'e': 'e', 'e': 'e', 'e': 'e', 'e': 'e', 'a': 'a', 'a': 'a', 'u': 'u', 'u': 'u', 'u': 'u',
            'o': 'o', 'o': 'o', 'i': 'i', 'i': 'i', 'c': 'c'}
    # Simple ASCII safe conversion
    result = ""
    for c in str(text):
        if ord(c) < 128:
            result += c
        elif c in 'éèêë':
            result += 'e'
        elif c in 'àâä':
            result += 'a'
        elif c in 'ùûü':
            result += 'u'
        elif c in 'ôö':
            result += 'o'
        elif c in 'îï':
            result += 'i'
        elif c == 'ç':
            result += 'c'
        elif c in 'ÉÈÊË':
            result += 'E'
        elif c in 'ÀÂÄ':
            result += 'A'
        elif c == 'Ç':
            result += 'C'
        elif c == '→':
            result += '->'
        elif c == '°':
            result += 'o'
        elif c == 'µ':
            result += 'u'
        elif c == '×':
            result += 'x'
        else:
            result += '?'
    return result


class PDF(FPDF):
    def header(self):
        # Top gradient bar
        self.set_fill_color(0, 20, 60)
        self.rect(0, 0, 210, 4, 'F')
        self.set_fill_color(0, 180, 255)
        self.rect(0, 4, 210, 1, 'F')
        self.set_fill_color(0, 255, 136)
        self.rect(0, 5, 210, 0.5, 'F')
        self.ln(8)
        
        # Logo area
        self.set_font("Arial", "B", 14)
        self.set_text_color(0, 60, 150)
        self.cell(0, 8, f"DM SMART LAB AI v{APP_VERSION}", 0, 0, "L")
        self.set_font("Arial", "", 8)
        self.set_text_color(100, 100, 100)
        self.cell(0, 8, datetime.now().strftime("%d/%m/%Y %H:%M:%S"), 0, 1, "R")
        
        # Subtitle
        self.set_font("Arial", "I", 7)
        self.set_text_color(120, 120, 120)
        self.cell(0, 4, "Systeme de Diagnostic Parasitologique par Intelligence Artificielle", 0, 1, "L")
        self.cell(0, 4, "INFSPM - Ouargla, Algerie", 0, 1, "L")
        
        # Line separator
        self.set_draw_color(0, 180, 255)
        self.set_line_width(0.5)
        self.line(10, self.get_y() + 2, 200, self.get_y() + 2)
        self.ln(6)

    def footer(self):
        self.set_y(-20)
        # Bottom gradient
        self.set_fill_color(0, 20, 60)
        self.rect(0, 282, 210, 15, 'F')
        self.set_y(-15)
        self.set_font("Arial", "I", 7)
        self.set_text_color(200, 200, 200)
        self.cell(0, 4, _sp("AVERTISSEMENT: Ce rapport est genere par IA - Validation par un professionnel de sante requise"), 0, 1, "C")
        self.set_font("Arial", "", 6)
        self.cell(0, 4, f"Page {self.page_no()}/{{nb}} | DM Smart Lab AI | Sebbag M.D.E & Ben Sghir M. | INFSPM Ouargla", 0, 0, "C")

    def section(self, title, color=(0, 60, 150)):
        self.set_fill_color(*color)
        self.set_text_color(255, 255, 255)
        self.set_font("Arial", "B", 10)
        self.cell(0, 8, f"  {_sp(title)}", 0, 1, "L", True)
        self.ln(2)
        self.set_text_color(0, 0, 0)

    def field(self, label, val):
        self.set_font("Arial", "B", 9)
        self.set_text_color(60, 60, 60)
        self.cell(55, 6, _sp(label), 0, 0)
        self.set_font("Arial", "", 9)
        self.set_text_color(0, 0, 0)
        self.cell(0, 6, _sp(str(val)), 0, 1)

    def add_separator(self):
        self.set_draw_color(200, 200, 200)
        self.set_line_width(0.2)
        self.line(10, self.get_y(), 200, self.get_y())
        self.ln(3)


def make_pdf(pat, lab, result, lbl):
    pdf = PDF()
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_auto_page_break(True, 25)

    # Title
    pdf.set_font("Arial", "B", 18)
    pdf.set_text_color(0, 40, 100)
    pdf.cell(0, 12, "RAPPORT D'ANALYSE PARASITOLOGIQUE", 0, 1, "C")

    # Reference ID
    rid = hashlib.md5(f"{pat.get('Nom', '')}{datetime.now().isoformat()}".encode()).hexdigest()[:10].upper()
    pdf.set_font("Arial", "", 8)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 5, f"Reference: DM-{rid} | Date: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 0, 1, "R")
    pdf.ln(3)

    # Patient section
    pdf.section("INFORMATIONS DU PATIENT")
    for k, v in pat.items():
        if v:
            pdf.field(f"{k}:", v)
    pdf.add_separator()

    # Lab section
    pdf.section("INFORMATIONS LABORATOIRE", (0, 100, 80))
    for k, v in lab.items():
        if v:
            pdf.field(f"{k}:", v)
    pdf.add_separator()

    # Result section
    cf = result.get("conf", 0)
    if lbl == "Negative":
        pdf.section("RESULTAT DE L'ANALYSE IA", (0, 150, 80))
    else:
        pdf.section("RESULTAT DE L'ANALYSE IA", (180, 0, 0))
    pdf.ln(2)

    # Result box
    if lbl == "Negative":
        pdf.set_fill_color(220, 255, 220)
        pdf.set_text_color(0, 100, 0)
    else:
        pdf.set_fill_color(255, 220, 220)
        pdf.set_text_color(180, 0, 0)

    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 12, f"  {_sp(lbl)} - Confiance: {cf}%", 1, 1, "C", True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(3)

    # Parasite details
    info = PARASITE_DB.get(lbl, PARASITE_DB["Negative"])

    pdf.set_font("Arial", "B", 9)
    pdf.cell(0, 6, "Nom Scientifique:", 0, 0)
    pdf.set_font("Arial", "I", 9)
    pdf.cell(0, 6, f"  {_sp(info.get('sci', 'N/A'))}", 0, 1)

    pdf.ln(2)
    pdf.set_font("Arial", "B", 9)
    pdf.cell(0, 6, "Morphologie:", 0, 1)
    pdf.set_font("Arial", "", 8)
    pdf.multi_cell(0, 5, _sp(info['morph'].get('fr', '')))

    pdf.ln(2)
    pdf.set_font("Arial", "B", 9)
    pdf.cell(0, 6, "Description:", 0, 1)
    pdf.set_font("Arial", "", 8)
    pdf.multi_cell(0, 5, _sp(info['desc'].get('fr', '')))

    pdf.ln(2)
    pdf.set_font("Arial", "B", 9)
    pdf.set_text_color(0, 100, 0)
    pdf.cell(0, 6, "Conseil Medical:", 0, 1)
    pdf.set_font("Arial", "", 8)
    pdf.set_text_color(0, 0, 0)
    pdf.multi_cell(0, 5, _sp(info['advice'].get('fr', '')))

    # Tests
    if info.get("tests"):
        pdf.ln(2)
        pdf.set_font("Arial", "B", 9)
        pdf.cell(0, 6, "Examens Complementaires Suggeres:", 0, 1)
        pdf.set_font("Arial", "", 8)
        for test in info["tests"]:
            pdf.cell(10, 5, "", 0, 0)
            pdf.cell(0, 5, f"- {_sp(test)}", 0, 1)

    # Reliability indicator
    pdf.ln(3)
    pdf.add_separator()
    rel_text = "FIABLE" if result.get("rel", False) else "A VERIFIER"
    rel_color = (0, 150, 0) if result.get("rel", False) else (200, 100, 0)
    pdf.set_font("Arial", "B", 10)
    pdf.set_text_color(*rel_color)
    pdf.cell(0, 8, f"Fiabilite: {rel_text} ({cf}%)", 0, 1, "C")
    pdf.set_text_color(0, 0, 0)

    # QR Code
    if HAS_QRCODE:
        try:
            qr = qrcode.QRCode(box_size=3, border=1)
            qr.add_data(f"DM-SmartLab|{lbl}|{cf}%|{rid}|{datetime.now().isoformat()}")
            qr.make(fit=True)
            qp = f"_qr_{rid}.png"
            qr.make_image().save(qp)
            pdf.image(qp, x=170, y=pdf.get_y() - 30, w=28)
            try:
                os.remove(qp)
            except:
                pass
        except:
            pass

    # Signatures section
    pdf.ln(8)
    pdf.section("SIGNATURES ET VALIDATION", (80, 80, 80))
    pdf.ln(5)
    pdf.set_font("Arial", "", 9)
    pdf.cell(95, 5, "Technicien 1: ___________________", 0, 0)
    pdf.cell(95, 5, "Technicien 2: ___________________", 0, 1)
    pdf.ln(8)
    pdf.cell(95, 5, "Date: ___/___/______", 0, 0)
    pdf.cell(95, 5, "Cachet du Laboratoire:", 0, 1)
    pdf.ln(10)
    pdf.cell(0, 5, "Validation Biologiste: ___________________", 0, 1)

    return bytes(pdf.output())


# ============================================
#  CSS - SPACE DARK THEME (Only Night Mode)
# ============================================
def apply_css():
    rtl = st.session_state.get("lang") == "ar"
    d = "rtl" if rtl else "ltr"
    
    # Space dark theme - ONLY dark mode
    bg = "#030614"
    cbg = "rgba(10,15,46,0.85)"
    tx = "#e0e8ff"
    pr = "#00f5ff"
    mu = "#6b7fa0"
    ac = "#ff00ff"
    ac2 = "#00ff88"
    sbg = "#020410"
    btn_bg = "linear-gradient(135deg,#00f5ff,#0066ff)"
    brd = "rgba(0,245,255,0.15)"
    ibg = "rgba(10,15,46,0.6)"
    template = "plotly_dark"
    
    st.markdown(f"""<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=Orbitron:wght@400;600;700;800;900&family=JetBrains+Mono:wght@400;600&family=Tajawal:wght@400;500;700;800&display=swap');
    
    html {{ direction: {d}; }}
    
    /* ====== ANIMATED SPACE BACKGROUND ====== */
    .stApp {{
        background: {bg} !important;
        color: {tx} !important;
        background-image: 
            radial-gradient(2px 2px at 20px 30px, rgba(0,245,255,0.3), transparent),
            radial-gradient(2px 2px at 40px 70px, rgba(255,0,255,0.2), transparent),
            radial-gradient(1px 1px at 90px 40px, rgba(0,255,136,0.3), transparent),
            radial-gradient(1px 1px at 130px 80px, rgba(0,245,255,0.2), transparent),
            radial-gradient(2px 2px at 160px 30px, rgba(255,0,255,0.15), transparent),
            radial-gradient(1px 1px at 200px 60px, rgba(0,255,136,0.2), transparent),
            radial-gradient(2px 2px at 60px 100px, rgba(0,102,255,0.2), transparent),
            radial-gradient(1px 1px at 180px 120px, rgba(255,105,180,0.15), transparent);
        background-size: 250px 150px;
        animation: sparkle 8s linear infinite;
    }}
    
    @keyframes sparkle {{
        0% {{ background-position: 0 0; }}
        100% {{ background-position: 250px 150px; }}
    }}
    
    /* ====== FLOATING PARTICLES OVERLAY ====== */
    .stApp::before {{
        content: '';
        position: fixed;
        top: 0;
        left: 0;
        width: 100%;
        height: 100%;
        pointer-events: none;
        z-index: 0;
        background: 
            radial-gradient(circle at 15% 25%, rgba(0,245,255,0.03) 0%, transparent 50%),
            radial-gradient(circle at 85% 75%, rgba(255,0,255,0.03) 0%, transparent 50%),
            radial-gradient(circle at 50% 50%, rgba(0,255,136,0.02) 0%, transparent 50%);
        animation: nebula 20s ease-in-out infinite;
    }}
    
    @keyframes nebula {{
        0%, 100% {{ opacity: 0.5; }}
        50% {{ opacity: 1; }}
    }}
    
    /* ====== SIDEBAR ====== */
    section[data-testid="stSidebar"] {{
        background: linear-gradient(180deg, {sbg} 0%, rgba(5,10,30,0.98) 100%) !important;
        border-right: 1px solid rgba(0,245,255,0.1) !important;
    }}
    section[data-testid="stSidebar"] * {{ color: {tx} !important; }}
    
    /* ====== TEXT COLORS ====== */
    .stApp p, .stApp span, .stApp label, .stApp div {{ color: {tx} !important; }}
    .stApp h1, .stApp h2, .stApp h3, .stApp h4 {{ color: {tx} !important; }}
    
    /* ====== INPUTS ====== */
    .stTextInput > div > div > input,
    .stTextArea > div > div > textarea,
    .stNumberInput > div > div > input {{
        background: {ibg} !important;
        color: {tx} !important;
        border: 1px solid {brd} !important;
        border-radius: 12px !important;
        transition: all 0.3s ease !important;
    }}
    .stTextInput > div > div > input:focus,
    .stTextArea > div > div > textarea:focus {{
        border-color: {pr} !important;
        box-shadow: 0 0 15px rgba(0,245,255,0.2) !important;
    }}
    
    /* ====== SELECTBOX ====== */
    .stSelectbox > div > div {{
        background: {ibg} !important;
        border: 1px solid {brd} !important;
        border-radius: 12px !important;
    }}
    
    /* ====== CARDS ====== */
    .dm-card {{
        background: {cbg};
        backdrop-filter: blur(20px);
        -webkit-backdrop-filter: blur(20px);
        border: 1px solid {brd};
        border-radius: 20px;
        padding: 24px;
        margin: 12px 0;
        transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
        color: {tx} !important;
        position: relative;
        overflow: hidden;
    }}
    .dm-card::before {{
        content: '';
        position: absolute;
        top: 0;
        left: -100%;
        width: 100%;
        height: 100%;
        background: linear-gradient(90deg, transparent, rgba(0,245,255,0.03), transparent);
        transition: left 0.6s ease;
    }}
    .dm-card:hover::before {{
        left: 100%;
    }}
    .dm-card:hover {{
        transform: translateY(-4px);
        box-shadow: 0 12px 40px rgba(0,245,255,0.1), 0 4px 15px rgba(255,0,255,0.05);
        border-color: rgba(0,245,255,0.3);
    }}
    .dm-card * {{ color: {tx} !important; }}
    .dm-card-cyan {{ border-left: 3px solid {pr}; }}
    .dm-card-green {{ border-left: 3px solid {ac2}; }}
    .dm-card-red {{ border-left: 3px solid #ff0040; }}
    .dm-card-purple {{ border-left: 3px solid #9933ff; }}
    
    /* ====== METRIC CARDS ====== */
    .dm-m {{
        background: {cbg};
        border: 1px solid {brd};
        border-radius: 18px;
        padding: 20px 12px;
        text-align: center;
        transition: all 0.3s ease;
        position: relative;
        overflow: hidden;
    }}
    .dm-m:hover {{
        border-color: {pr};
        box-shadow: 0 0 20px rgba(0,245,255,0.1);
    }}
    .dm-m-i {{ font-size: 1.6rem; }}
    .dm-m-v {{
        font-size: 1.8rem;
        font-weight: 800;
        font-family: 'JetBrains Mono', monospace !important;
        background: linear-gradient(135deg, {pr}, {ac});
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        animation: glow-text 3s ease-in-out infinite;
    }}
    @keyframes glow-text {{
        0%, 100% {{ filter: brightness(1); }}
        50% {{ filter: brightness(1.3); }}
    }}
    .dm-m-l {{
        font-size: .7rem;
        font-weight: 600;
        color: {mu} !important;
        -webkit-text-fill-color: {mu} !important;
        text-transform: uppercase;
        letter-spacing: .1em;
        margin-top: 6px;
    }}
    
    /* ====== BUTTONS ====== */
    div.stButton > button {{
        background: {btn_bg} !important;
        color: white !important;
        border: none !important;
        border-radius: 14px !important;
        padding: 12px 28px !important;
        font-weight: 700 !important;
        font-size: 0.85rem !important;
        letter-spacing: 0.03em !important;
        transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1) !important;
        position: relative !important;
        overflow: hidden !important;
    }}
    div.stButton > button::before {{
        content: '' !important;
        position: absolute !important;
        top: 0 !important;
        left: -100% !important;
        width: 100% !important;
        height: 100% !important;
        background: linear-gradient(90deg, transparent, rgba(255,255,255,0.2), transparent) !important;
        transition: left 0.5s ease !important;
    }}
    div.stButton > button:hover::before {{
        left: 100% !important;
    }}
    div.stButton > button:hover {{
        transform: translateY(-3px) !important;
        box-shadow: 0 8px 25px rgba(0,245,255,0.3), 0 0 40px rgba(0,102,255,0.15) !important;
    }}
    div.stButton > button * {{
        color: white !important;
        -webkit-text-fill-color: white !important;
    }}
    
    /* ====== NEON TITLE ====== */
    .dm-nt {{
        font-family: 'Orbitron', sans-serif;
        font-weight: 900;
        background: linear-gradient(135deg, {pr}, {ac}, {ac2});
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-size: 200% auto;
        animation: gradient-shift 4s ease infinite;
    }}
    @keyframes gradient-shift {{
        0% {{ background-position: 0% center; }}
        50% {{ background-position: 100% center; }}
        100% {{ background-position: 0% center; }}
    }}
    
    /* ====== CHAT BUBBLES ====== */
    .dm-ch {{
        padding: 14px 18px;
        border-radius: 18px;
        margin: 8px 0;
        max-width: 85%;
        font-size: .9rem;
        line-height: 1.6;
        animation: fadeInUp 0.3s ease;
    }}
    @keyframes fadeInUp {{
        from {{ opacity: 0; transform: translateY(10px); }}
        to {{ opacity: 1; transform: translateY(0); }}
    }}
    .dm-cu {{
        background: linear-gradient(135deg, #0066ff, #0044cc) !important;
        color: white !important;
        margin-left: auto;
        border-bottom-right-radius: 4px;
    }}
    .dm-cu * {{ color: white !important; -webkit-text-fill-color: white !important; }}
    .dm-cb {{
        background: {cbg};
        border: 1px solid {brd};
        border-bottom-left-radius: 4px;
    }}
    
    /* ====== HEADINGS ====== */
    h1 {{ font-family: 'Orbitron', sans-serif !important; }}
    
    /* ====== RTL Support ====== */
    {"body, p, span, div, label { font-family: 'Tajawal', sans-serif !important; }" if rtl else ""}
    
    /* ====== LOGO CONTAINER ====== */
    .dm-logo {{
        text-align: center;
        padding: 16px;
        background: linear-gradient(135deg, rgba(0,245,255,0.05), rgba(255,0,255,0.05));
        border-radius: 24px;
        border: 1px solid {brd};
        margin-bottom: 12px;
        position: relative;
        overflow: hidden;
    }}
    .dm-logo::after {{
        content: '';
        position: absolute;
        top: -50%;
        left: -50%;
        width: 200%;
        height: 200%;
        background: conic-gradient(from 0deg, transparent, rgba(0,245,255,0.05), transparent, rgba(255,0,255,0.05), transparent);
        animation: logo-rotate 10s linear infinite;
    }}
    @keyframes logo-rotate {{
        100% {{ transform: rotate(360deg); }}
    }}
    
    /* ====== EXPANDER ====== */
    .streamlit-expanderHeader {{
        background: rgba(10,15,46,0.5) !important;
        border-radius: 12px !important;
        border: 1px solid {brd} !important;
    }}
    
    /* ====== TABS ====== */
    .stTabs [data-baseweb="tab-list"] {{
        gap: 4px;
        background: rgba(10,15,46,0.5);
        border-radius: 14px;
        padding: 4px;
    }}
    .stTabs [data-baseweb="tab"] {{
        border-radius: 10px;
        color: {tx} !important;
    }}
    .stTabs [aria-selected="true"] {{
        background: linear-gradient(135deg, rgba(0,245,255,0.2), rgba(0,102,255,0.2)) !important;
    }}
    
    /* ====== PROGRESS BAR ====== */
    .stProgress > div > div > div > div {{
        background: linear-gradient(90deg, {pr}, {ac}, {ac2}) !important;
        border-radius: 10px !important;
        animation: progress-glow 2s ease infinite;
    }}
    @keyframes progress-glow {{
        0%, 100% {{ box-shadow: 0 0 5px rgba(0,245,255,0.3); }}
        50% {{ box-shadow: 0 0 15px rgba(0,245,255,0.5); }}
    }}
    
    /* ====== DATAFRAME ====== */
    .stDataFrame {{
        border-radius: 14px !important;
        overflow: hidden !important;
    }}
    
    /* ====== RADIO ====== */
    .stRadio > div {{
        gap: 4px;
    }}
    
    /* ====== SCROLLBAR ====== */
    ::-webkit-scrollbar {{
        width: 6px;
        height: 6px;
    }}
    ::-webkit-scrollbar-track {{
        background: {bg};
    }}
    ::-webkit-scrollbar-thumb {{
        background: linear-gradient(180deg, {pr}, {ac});
        border-radius: 10px;
    }}
    
    </style>""", unsafe_allow_html=True)
    return template


plot_template = apply_css()


# ============================================
#  ANIMATED LOGO
# ============================================
def render_logo():
    st.markdown("""<div class="dm-logo">
    <svg width="75" height="75" viewBox="0 0 100 100" xmlns="http://www.w3.org/2000/svg">
    <defs>
        <linearGradient id="g1" x1="0%" y1="0%" x2="100%" y2="100%">
            <stop offset="0%" style="stop-color:#00f5ff">
                <animate attributeName="stop-color" values="#00f5ff;#ff00ff;#00ff88;#00f5ff" dur="4s" repeatCount="indefinite"/>
            </stop>
            <stop offset="50%" style="stop-color:#ff00ff">
                <animate attributeName="stop-color" values="#ff00ff;#00ff88;#00f5ff;#ff00ff" dur="4s" repeatCount="indefinite"/>
            </stop>
            <stop offset="100%" style="stop-color:#00ff88">
                <animate attributeName="stop-color" values="#00ff88;#00f5ff;#ff00ff;#00ff88" dur="4s" repeatCount="indefinite"/>
            </stop>
        </linearGradient>
        <filter id="gl">
            <feGaussianBlur stdDeviation="2" result="b"/>
            <feMerge><feMergeNode in="b"/><feMergeNode in="SourceGraphic"/></feMerge>
        </filter>
    </defs>
    <circle cx="50" cy="50" r="45" fill="none" stroke="url(#g1)" stroke-width="2.5" filter="url(#gl)" opacity=".8">
        <animateTransform attributeName="transform" type="rotate" values="0 50 50;360 50 50" dur="20s" repeatCount="indefinite"/>
    </circle>
    <circle cx="50" cy="50" r="38" fill="none" stroke="url(#g1)" stroke-width="1" opacity=".3">
        <animateTransform attributeName="transform" type="rotate" values="360 50 50;0 50 50" dur="15s" repeatCount="indefinite"/>
    </circle>
    <circle cx="50" cy="50" r="30" fill="none" stroke="url(#g1)" stroke-width="0.5" opacity=".2"/>
    <text x="50" y="42" text-anchor="middle" font-family="Orbitron,sans-serif" font-size="14" font-weight="900" fill="url(#g1)" filter="url(#gl)">DM</text>
    <text x="50" y="58" text-anchor="middle" font-family="Orbitron,sans-serif" font-size="8" font-weight="600" fill="url(#g1)">SMART LAB</text>
    <text x="50" y="68" text-anchor="middle" font-family="Orbitron,sans-serif" font-size="7" font-weight="400" fill="#00f5ff" opacity=".6">AI</text>
    <!-- DNA Helix dots -->
    <circle cx="25" cy="35" r="2" fill="#00f5ff" opacity=".4"><animate attributeName="cy" values="35;65;35" dur="3s" repeatCount="indefinite"/></circle>
    <circle cx="75" cy="65" r="2" fill="#ff00ff" opacity=".4"><animate attributeName="cy" values="65;35;65" dur="3s" repeatCount="indefinite"/></circle>
    <circle cx="30" cy="50" r="1.5" fill="#00ff88" opacity=".3"><animate attributeName="cx" values="30;70;30" dur="4s" repeatCount="indefinite"/></circle>
    </svg>
    <h3 style="font-family:Orbitron,sans-serif;margin:6px 0;background:linear-gradient(135deg,#00f5ff,#ff00ff,#00ff88);-webkit-background-clip:text;-webkit-text-fill-color:transparent;font-size:1.1rem;position:relative;z-index:1;">DM SMART LAB AI</h3>
    <p style="font-size:.55rem;opacity:.35;letter-spacing:.35em;text-transform:uppercase;margin:0;position:relative;z-index:1;">v7.5 Space Edition</p>
    </div>""", unsafe_allow_html=True)


# ============================================
#  SIDEBAR LOGO
# ============================================
with st.sidebar:
    render_logo()

# ============================================
#  LOGIN PAGE
# ============================================

if not st.session_state.logged_in:
    lc1, lc2, lc3 = st.columns([1, 2, 1])
    with lc2:
        # Language selector
        ll = st.selectbox(
            t("language"),
            ["FR 🇫🇷 Français", "AR 🇸🇦 العربية", "EN 🇬🇧 English"],
            label_visibility="collapsed",
            key="login_lang"
        )
        st.session_state.lang = "fr" if "FR" in ll else ("ar" if "AR" in ll else "en")

        render_logo()

        st.markdown(f"""<div class='dm-card dm-card-cyan' style='text-align:center;'>
        <div style='font-size:3.5rem;margin-bottom:10px;animation: pulse 2s ease-in-out infinite;'>
            🔐
        </div>
        <h2 class='dm-nt'>{t('login_title')}</h2>
        <p style='opacity:.4;font-size:.85rem;'>{t('login_subtitle')}</p>
        </div>
        <style>
        @keyframes pulse {{
            0%, 100% {{ transform: scale(1); }}
            50% {{ transform: scale(1.1); }}
        }}
        </style>
        """, unsafe_allow_html=True)

        with st.form("login_form"):
            u = st.text_input(
                f"{t('username')} *",
                placeholder="admin / dhia / demo / tech1 / tech2",
                key="login_username"
            )
            p = st.text_input(f"{t('password')} *", type="password", key="login_password")
            
            login_btn = st.form_submit_button(
                f"🔐 {t('connect')}",
                use_container_width=True,
                type="primary"
            )
            
            if login_btn:
                if not u.strip():
                    st.error(f"❌ {t('username')} {t('name_required')}")
                else:
                    with st.spinner(f"🔄 {t('connect')}..."):
                        result = db_login(u.strip(), p)
                        
                        if result is None:
                            st.error("❌ User not found")
                        elif isinstance(result, dict) and "error" in result:
                            if result["error"] == "locked":
                                st.error("🔒 Account locked. Try again later.")
                            else:
                                left = result.get("left", 0)
                                st.error(f"❌ Wrong password. {left} attempts left")
                        else:
                            st.session_state.logged_in = True
                            st.session_state.user_id = result["id"]
                            st.session_state.user_name = result["username"]
                            st.session_state.user_role = result["role"]
                            st.session_state.user_full_name = result["full_name"]
                            db_log(result["id"], result["username"], "Login")
                            st.success("✅ Login successful!")
                            st.balloons()
                            st.rerun()

        st.markdown(f"""<div style='text-align:center;opacity:.2;font-size:.65rem;margin-top:20px;font-family:monospace;'>
        admin/admin2026 | dhia/dhia2026 | tech1/tech2026 | demo/demo123
        </div>""", unsafe_allow_html=True)
    
    st.stop()

# ============================================
#  SIDEBAR NAVIGATION
# ============================================
with st.sidebar:
    ri = ROLES.get(st.session_state.user_role, ROLES["viewer"])
    st.markdown(f"### {ri['icon']} {st.session_state.user_full_name}")
    st.caption(f"@{st.session_state.user_name} | {tl(ri['label'])}")

    tips = TIPS.get(st.session_state.lang, TIPS["fr"])
    st.info(f"**{t('daily_tip')}:**\n\n{tips[datetime.now().timetuple().tm_yday % len(tips)]}")

    st.markdown("---")
    st.markdown(f"#### {t('language')}")
    lc = st.radio("lang_select", ["FR Francais", "AR العربية", "EN English"], label_visibility="collapsed",
                  index=["fr", "ar", "en"].index(st.session_state.lang))
    nl = "fr" if "FR" in lc else ("ar" if "AR" in lc else "en")
    if nl != st.session_state.lang:
        st.session_state.lang = nl
        st.rerun()

    st.markdown("---")

    # Navigation
    navs = [
        f"🏠 {t('home')}",
        f"🔬 {t('scan')}",
        f"📘 {t('encyclopedia')}",
        f"📊 {t('dashboard')}",
        f"🧠 {t('quiz')}",
        f"💬 {t('chatbot')}",
        f"🔄 {t('compare')}",
    ]
    keys = ["home", "scan", "enc", "dash", "quiz", "chat", "cmp"]
    if has_role(3):
        navs.append(f"⚙️ {t('admin')}")
        keys.append("admin")
    navs.append(f"ℹ️ {t('about')}")
    keys.append("about")

    menu = st.radio("Navigation", navs, label_visibility="collapsed")

    st.markdown("---")
    if st.button(f"🚪 {t('logout')}", use_container_width=True):
        db_log(st.session_state.user_id, st.session_state.user_name, "Logout")
        for k in DEFAULTS:
            st.session_state[k] = DEFAULTS[k]
        st.rerun()

pg = dict(zip(navs, keys)).get(menu, "home")

# Render voice player at top of page
render_voice_player()


# ════════════════════════════════════════════
#  PAGE: HOME
# ════════════════════════════════════════════
if pg == "home":
    st.markdown(f"""<h1 style='font-family:Orbitron,sans-serif;'>
    <span class='dm-nt'>👋 {get_greeting()}, {st.session_state.user_full_name}!</span>
    </h1>""", unsafe_allow_html=True)

    st.markdown(f"""<div class='dm-card dm-card-cyan'>
    <h2 class='dm-nt'>🧬 DM SMART LAB AI</h2>
    <h4 style='opacity:.6;'>{t('where_science')}</h4>
    <p style='opacity:.4;font-size:.85rem;'>{t('system_desc')}</p>
    <p style='margin-top: 15px; padding: 10px; background: rgba(0,255,136,0.1); border-radius: 8px; border-left: 3px solid #00ff88;'>
    <b>✨ v{APP_VERSION} - Professional Edition</b><br>
    Diagnostic Parasitologique par Intelligenc Artificielle<br>
    <small>INFSPM Ouargla, Algérie 🇩🇿</small>
    </p>
    </div>""", unsafe_allow_html=True)

    st.markdown("---")

    # Voice buttons
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        if st.button(f"🎙️ {t('welcome_btn')}", use_container_width=True, type="primary", key="home_welcome"):
            speak(t("voice_welcome"))
            st.rerun()
    with col2:
        if st.button(f"🤖 {t('intro_btn')}", use_container_width=True, type="primary", key="home_intro"):
            speak(t("voice_intro"))
            st.rerun()
    with col3:
        if st.button(f"🔇 {t('stop_voice')}", use_container_width=True, key="home_stop"):
            stop_speech()
    with col4:
        if st.button("🌍 Multilingual", use_container_width=True, disabled=True):
            pass

    st.markdown("---")
    st.markdown(f"### 📊 {t('quick_overview')}")
    
    # Get stats
    stats = db_stats(st.session_state.user_id)
    
    metric_cols = st.columns(5)
    metrics = [
        ("🔬", stats["total"], t("total_analyses")),
        ("✅", stats["reliable"], t("reliable")),
        ("⚠️", stats["verify"], t("to_verify")),
        ("🦠", stats["top"], t("most_frequent")),
        ("📈", f"{stats['avg']}%", t("avg_confidence"))
    ]
    
    for col, (ic, v, lb) in zip(metric_cols, metrics):
        with col:
            st.markdown(f"""<div class='dm-m'>
            <span class='dm-m-i'>{ic}</span>
            <div class='dm-m-v'>{v}</div>
            <div class='dm-m-l'>{lb}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown(f"### 🎯 Quick Access")
    
    quick_cols = st.columns(4)
    quick_actions = [
        ("🔬", t("scan"), "Analyze parasite samples"),
        ("🧠", t("quiz"), "Test your knowledge"),
        ("💬", t("chatbot"), "Ask DM Bot"),
        ("📊", t("dashboard"), "View analytics")
    ]
    
    for col, (icon, title, desc) in zip(quick_cols, quick_actions):
        with col:
            st.markdown(f"""<div style='background: rgba(0,245,255,0.1); border: 1px solid rgba(0,245,255,0.2); 
                            padding: 15px; border-radius: 12px; text-align: center; cursor: pointer;
                            transition: all 0.3s; hover:box-shadow: 0 0 20px rgba(0,245,255,0.3);'>
            <div style='font-size: 2rem;'>{icon}</div>
            <p style='font-weight: bold; margin: 8px 0;'>{title}</p>
            <p style='font-size: 0.75rem; opacity: 0.6;'>{desc}</p>
            </div>""", unsafe_allow_html=True)

# ════════════════════════════════════════════
#  PAGE: SCAN
# ════════════════════════════════════════════
elif pg == "scan":
    st.title(f"🔬 {t('scan')}")
    mdl, mn, mt = load_model()
    if mn:
        st.sidebar.success(f"🧠 Model: {mn}")
    else:
        st.sidebar.info(f"🧠 {t('demo_mode')}")

    st.markdown(f"### 📋 1. {t('patient_info')}")
    ca, cb = st.columns(2)
    pn = ca.text_input(f"{t('patient_name')} *")
    pf = cb.text_input(t("patient_firstname"))
    c1, c2, c3, c4 = st.columns(4)
    pa = c1.number_input(t("age"), 0, 120, 30)
    ps = c2.selectbox(t("sex"), [t("male"), t("female")])
    pw = c3.number_input(t("weight"), 0, 300, 70)
    pst = c4.selectbox(t("sample_type"), SAMPLES.get(st.session_state.lang, SAMPLES["fr"]))

    st.markdown(f"### 🔬 2. {t('lab_info')}")
    la, lb2, lc2 = st.columns(3)
    t1 = la.text_input(f"{t('technician')} 1", value=st.session_state.user_full_name)
    t2 = lb2.text_input(f"{t('technician')} 2")
    lm = lc2.selectbox(t("microscope"), MICROSCOPE_TYPES)
    ld, le = st.columns(2)
    mg = ld.selectbox(t("magnification"), MAGNIFICATIONS, index=3)
    pt = le.selectbox(t("preparation"), PREPARATION_TYPES)
    nt = st.text_area(t("notes"), height=80)

    st.markdown("---")
    st.markdown(f"### 📸 3. {t('image_capture')}")
    src = st.radio("source", [t("take_photo"), t("upload_file")], horizontal=True, label_visibility="collapsed")
    img = None
    ih = None

    if t("take_photo") in src:
        st.info(f"📷 {t('camera_hint')}")
        cp = st.camera_input("camera")
        if cp:
            img = Image.open(cp).convert("RGB")
            ih = hashlib.md5(cp.getvalue()).hexdigest()
    else:
        uf = st.file_uploader(t("upload_file"), type=["jpg", "jpeg", "png", "bmp", "tiff"])
        if uf:
            img = Image.open(uf).convert("RGB")
            ih = hashlib.md5(uf.getvalue()).hexdigest()

    if img:
        if not pn.strip():
            st.error(t("name_required"))
            st.stop()

        if st.session_state.get("_ih") != ih:
            st.session_state._ih = ih
            st.session_state.demo_seed = random.randint(0, 999999)
            st.session_state.heatmap_seed = random.randint(0, 999999)

        ci, cr = st.columns(2)
        with ci:
            with st.expander("🎛️ Zoom / Adjust"):
                z = st.slider("Zoom", 1.0, 5.0, 1.0, 0.25)
                br = st.slider("Brightness", 0.5, 2.0, 1.0, 0.1)
                co = st.slider("Contrast", 0.5, 2.0, 1.0, 0.1)
                sa = st.slider("Saturation", 0.0, 2.0, 1.0, 0.1)
            adj = adjust(img, br, co, sa)
            if z > 1:
                adj = zoom_img(adj, z)

            tabs = st.tabs(["📷 Original", "🔥 Thermal", "📐 Edges", "✨ Enhanced", "🔄 Negative", "🏔️ Emboss", "🎯 Heatmap"])
            with tabs[0]:
                st.image(adj, use_container_width=True)
            with tabs[1]:
                st.image(thermal(adj), use_container_width=True)
            with tabs[2]:
                st.image(edges_filter(adj), use_container_width=True)
            with tabs[3]:
                st.image(enhanced_filter(adj), use_container_width=True)
            with tabs[4]:
                st.image(negative_filter(adj), use_container_width=True)
            with tabs[5]:
                st.image(emboss_filter(adj), use_container_width=True)
            with tabs[6]:
                st.image(gen_heatmap(img, st.session_state.heatmap_seed), use_container_width=True)

        with cr:
            st.markdown(f"### 🧠 {t('result')}")
            with st.spinner("Analyzing..."):
                pg2 = st.progress(0)
                for i in range(100):
                    time.sleep(0.003)
                    pg2.progress(i + 1)
                res = predict(mdl, mt, img, st.session_state.demo_seed)

            lb_result = res["label"]
            cf = res["conf"]
            rc = risk_color(res["risk"])
            info = PARASITE_DB.get(lb_result, PARASITE_DB["Negative"])

            if not res["rel"]:
                st.warning(t("low_conf_warn"))
            if res["demo"]:
                st.info(t("demo_mode"))

            st.markdown(f"""<div class='dm-card' style='border-left:4px solid {rc};'>
            <div style='display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:10px;'>
            <div>
                <h2 style='color:{rc}!important;-webkit-text-fill-color:{rc}!important;margin:0;font-family:Orbitron;'>{lb_result}</h2>
                <p style='opacity:.4;font-style:italic;'>{info['sci']}</p>
            </div>
            <div style='text-align:center;'>
                <div style='font-size:2.8rem;font-weight:900;font-family:JetBrains Mono;color:{rc}!important;-webkit-text-fill-color:{rc}!important;'>{cf}%</div>
                <div style='font-size:.7rem;opacity:.4;'>{t("confidence")}</div>
            </div>
            </div>
            <hr style='opacity:.1;margin:14px 0;'>
            <p><b>🔬 {t("morphology")}:</b><br>{tl(info['morph'])}</p>
            <p><b>⚠️ {t("risk")}:</b> <span style='color:{rc}!important;-webkit-text-fill-color:{rc}!important;font-weight:700;'>{tl(info['risk_d'])}</span></p>
            <div style='background:rgba(0,255,136,.06);padding:14px;border-radius:12px;margin:10px 0;border:1px solid rgba(0,255,136,.1);'>
                <b>💡 {t("advice")}:</b><br>{tl(info['advice'])}
            </div>
            <div style='background:rgba(0,100,255,.06);padding:14px;border-radius:12px;font-style:italic;border:1px solid rgba(0,100,255,.1);'>
                🤖 {tl(info['funny'])}
            </div>
            </div>""", unsafe_allow_html=True)

            vc1, vc2 = st.columns(2)
            with vc1:
                if st.button(f"🔊 {t('listen')}", use_container_width=True):
                    speak(f"{lb_result}. {tl(info['funny'])}")
                    st.rerun()
            with vc2:
                if st.button(f"🔇 {t('stop_voice')}", key="sv2", use_container_width=True):
                    stop_speech()

            if info.get("tests"):
                with st.expander(f"🩺 {t('extra_tests')}"):
                    for x in info["tests"]:
                        st.markdown(f"- {x}")

            dk = tl(info.get("keys", {}))
            if dk and dk not in ["N/A", "غير متوفر"]:
                with st.expander(f"🔑 {t('diagnostic_keys')}"):
                    st.markdown(dk)

            cy = tl(info.get("cycle", {}))
            if cy and cy not in ["N/A", "غير متوفر"]:
                with st.expander(f"🔄 {t('lifecycle')}"):
                    st.markdown(f"**{cy}**")

            if res["preds"] and HAS_PLOTLY:
                with st.expander(f"📊 {t('all_probabilities')}"):
                    sp = dict(sorted(res["preds"].items(), key=lambda x: x[1], reverse=True))
                    fig = px.bar(x=list(sp.values()), y=list(sp.keys()), orientation='h',
                                 color=list(sp.values()), color_continuous_scale='RdYlGn_r')
                    fig.update_layout(height=220, template=plot_template,
                                      margin=dict(l=20, r=20, t=10, b=20), showlegend=False)
                    st.plotly_chart(fig, use_container_width=True)

        st.markdown("---")
        a1, a2, a3 = st.columns(3)
        with a1:
            try:
                pdf = make_pdf(
                    {"Nom": pn, "Prenom": pf, "Age": str(pa), "Sexe": ps, "Poids": str(pw), "Echantillon": pst},
                    {"Microscope": lm, "Grossissement": mg, "Preparation": pt, "Tech1": t1, "Tech2": t2, "Notes": nt},
                    res, lb_result)
                st.download_button(f"📥 {t('download_pdf')}", pdf,
                                   f"Rapport_{pn}_{datetime.now().strftime('%Y%m%d')}.pdf",
                                   "application/pdf", use_container_width=True)
            except Exception as e:
                st.error(f"PDF Error: {e}")
        with a2:
            if has_role(2) and st.button(f"💾 {t('save_db')}", use_container_width=True):
                aid = db_save_analysis(st.session_state.user_id, {
                    "pn": pn, "pf": pf, "pa": pa, "ps": ps, "pw": pw,
                    "st": pst, "mt": lm, "mg": mg, "pt": pt, "t1": t1, "t2": t2, "nt": nt,
                    "label": lb_result, "conf": cf, "risk": res["risk"],
                    "rel": 1 if res["rel"] else 0,
                    "preds": res["preds"], "hash": ih, "demo": 1 if res["demo"] else 0
                })
                db_log(st.session_state.user_id, st.session_state.user_name, "Analysis", f"ID:{aid}")
                st.success(t("saved_ok"))
        with a3:
            if st.button(f"🔄 {t('new_analysis')}", use_container_width=True):
                st.session_state.demo_seed = None
                st.session_state._ih = None
                st.rerun()


# ════════════════════════════════════════════
#  PAGE: ENCYCLOPEDIA
# ════════════════════════════════════════════
elif pg == "enc":
    st.title(f"📘 {t('encyclopedia')}")
    sr = st.text_input(f"🔍 {t('search')}")
    st.markdown("---")
    found = False
    for nm, d in PARASITE_DB.items():
        if nm == "Negative":
            continue
        if sr.strip() and sr.lower() not in (nm + " " + d["sci"]).lower():
            continue
        found = True
        rc = risk_color(d["risk"])
        with st.expander(f"{d['icon']} {nm} -- *{d['sci']}* | {tl(d['risk_d'])}", expanded=not sr.strip()):
            ci, cv = st.columns([2.5, 1])
            with ci:
                st.markdown(f"""<div class='dm-card' style='border-left:3px solid {rc};'>
                <h4 style='color:{rc}!important;-webkit-text-fill-color:{rc}!important;font-family:Orbitron;'>{d['sci']}</h4>
                <p><b>🔬 {t("morphology")}:</b><br>{tl(d['morph'])}</p>
                <p><b>📖 {t("description")}:</b><br>{tl(d['desc'])}</p>
                <p><b>⚠️ {t("risk")}:</b> <span style='color:{rc}!important;-webkit-text-fill-color:{rc}!important;font-weight:700;'>{tl(d['risk_d'])}</span></p>
                <div style='background:rgba(0,255,136,.06);padding:12px;border-radius:10px;margin:8px 0;'><b>💡:</b> {tl(d['advice'])}</div>
                <div style='background:rgba(0,100,255,.06);padding:12px;border-radius:10px;font-style:italic;'>🤖 {tl(d['funny'])}</div>
                </div>""", unsafe_allow_html=True)
                cy = tl(d.get("cycle", {}))
                if cy and cy not in ["N/A", "غير متوفر"]:
                    st.markdown(f"**🔄 {t('lifecycle')}:** {cy}")
                dk = tl(d.get("keys", {}))
                if dk:
                    st.markdown(f"**🔑 {t('diagnostic_keys')}:**\n{dk}")
                if d.get("tests"):
                    st.markdown(f"**🩺 {t('extra_tests')}:** {', '.join(d['tests'])}")
            with cv:
                rp = risk_pct(d["risk"])
                if rp > 0:
                    st.progress(rp / 100, text=f"{rp}%")
                st.markdown(f'<div style="text-align:center;font-size:4rem;">{d["icon"]}</div>', unsafe_allow_html=True)
                if st.button(f"🔊 {t('listen')}", key=f"ev_{nm}"):
                    speak(f"{nm}. {tl(d['desc'])}")
                    st.rerun()
    if sr.strip() and not found:
        st.warning(t("no_results"))


# ════════════════════════════════════════════
#  PAGE: DASHBOARD
# ════════════════════════════════════════════
elif pg == "dash":
    st.title(f"📊 {t('dashboard')}")
    
    # Get data
    s = db_stats() if has_role(3) else db_stats(st.session_state.user_id)
    an = db_analyses() if has_role(3) else db_analyses(st.session_state.user_id)

    # Metrics
    metric_cols = st.columns(5)
    metrics_data = [
        ("🔬", s["total"], t("total_analyses")),
        ("✅", s["reliable"], t("reliable")),
        ("⚠️", s["verify"], t("to_verify")),
        ("🦠", s["top"], t("most_frequent")),
        ("📈", f"{s['avg']}%", t("avg_confidence"))
    ]
    
    for col, (ic, v, lb) in zip(metric_cols, metrics_data):
        with col:
            st.markdown(f"""<div class='dm-m'>
            <span class='dm-m-i'>{ic}</span>
            <div class='dm-m-v'>{v}</div>
            <div class='dm-m-l'>{lb}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("---")

    if s["total"] > 0 and an:
        df = pd.DataFrame(an)
        
        chart_cols = st.columns(2)
        
        with chart_cols[0]:
            st.markdown(f"#### {t('parasite_distribution')}")
            if "parasite_detected" in df.columns:
                pc = df["parasite_detected"].value_counts().head(10)
                fig = px.pie(
                    values=pc.values,
                    names=pc.index,
                    hole=0.4,
                    color_discrete_sequence=px.colors.sequential.Plasma
                )
                fig.update_layout(height=350, template='plotly_dark', 
                                 margin=dict(l=20, r=20, t=20, b=20))
                st.plotly_chart(fig, use_container_width=True)

        with chart_cols[1]:
            st.markdown(f"#### {t('confidence_levels')}")
            if "confidence" in df.columns:
                fig = px.histogram(df, x="confidence", nbins=20,
                                  color_discrete_sequence=["#00f5ff"])
                fig.update_layout(height=350, template='plotly_dark',
                                 margin=dict(l=20, r=20, t=20, b=20))
                st.plotly_chart(fig, use_container_width=True)

        # Trends
        tr = db_trends(30)
        if tr:
            st.markdown(f"#### {t('trends')}")
            tr_df = pd.DataFrame(tr)
            fig = px.line(tr_df, x="day", y="count", color="parasite_detected",
                         markers=True, title="30-Day Trend")
            fig.update_layout(height=300, template='plotly_dark',
                             margin=dict(l=20, r=20, t=40, b=20))
            st.plotly_chart(fig, use_container_width=True)

        # History table
        st.markdown(f"### 📋 {t('history')}")
        display_cols = [c for c in ["id", "analysis_date", "patient_name", 
                                   "parasite_detected", "confidence", "risk_level", 
                                   "analyst", "validated"] if c in df.columns]
        st.dataframe(df[display_cols] if display_cols else df, use_container_width=True, 
                    height=400)

        # Validation
        if has_role(2) and "validated" in df.columns:
            uv = df[df["validated"] == 0]
            if not uv.empty:
                st.markdown("---")
                st.markdown("### ✔️ Pending Validations")
                vi = st.selectbox("Select Analysis ID:", uv["id"].tolist(), 
                                 key="dash_validate_id")
                if st.button(f"✅ Validate #{vi}", key="dash_validate_btn"):
                    db_validate(vi, st.session_state.user_full_name)
                    st.success(f"✅ Validated #{vi}")
                    st.rerun()

        # Export
        st.markdown("---")
        export_cols = st.columns(3)
        with export_cols[0]:
            st.download_button(f"⬇️ {t('export_csv')}", 
                              df.to_csv(index=False).encode('utf-8-sig'),
                              f"analyses_{datetime.now().strftime('%Y%m%d')}.csv",
                              "text/csv", use_container_width=True, 
                              key="dash_csv")
        with export_cols[1]:
            st.download_button(f"⬇️ {t('export_json')}",
                              df.to_json(orient='records', force_ascii=False).encode(),
                              f"analyses_{datetime.now().strftime('%Y%m%d')}.json",
                              "application/json", use_container_width=True,
                              key="dash_json")
        with export_cols[2]:
            try:
                from openpyxl import Workbook
                xlsx_buffer = io.BytesIO()
                df.to_excel(xlsx_buffer, index=False, engine='openpyxl')
                xlsx_buffer.seek(0)
                st.download_button(f"⬇️ Excel",
                                  xlsx_buffer.getvalue(),
                                  f"analyses_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                  use_container_width=True, key="dash_xlsx")
            except:
                st.info("Excel export requires openpyxl")
    else:
        st.info(f"ℹ️ {t('no_data')}")

# ════════════════════════════════════════════
#  PAGE: QUIZ (Fixed & Enhanced)
# ════════════════════════════════════════════
elif pg == "quiz":
    st.title(f"🧠 {t('quiz')}")

    # Initialize quiz state properly
    if "quiz_state" not in st.session_state:
        st.session_state.quiz_state = {
            "current": 0, "score": 0, "answered": [],
            "active": False, "order": [], "wrong": [],
            "total_q": 0, "finished": False, "selected_answer": None
        }

    qs = st.session_state.quiz_state

    # Leaderboard
    with st.expander(f"🏆 {t('leaderboard')}"):
        lb_data = db_leaderboard()
        if lb_data:
            for i, e in enumerate(lb_data[:10]):
                medal = "🥇" if i == 0 else "🥈" if i == 1 else "🥉" if i == 2 else f"**#{i + 1}**"
                st.markdown(f"{medal} **{e['username']}** — {e['score']}/{e['total_questions']} ({e['percentage']:.0f}%)")
        else:
            st.info(t("no_data"))

    # Quiz not started
# Quiz not started
    if not qs.get("active", False) and not qs.get("finished", False):
        st.markdown(f"""<div class='dm-card dm-card-cyan' style='text-align:center;'>
        <div style='font-size:4rem;margin-bottom:10px;'>🧠</div>
        <h3 class='dm-nt'>{ {"fr":"Testez vos connaissances en parasitologie !","ar":"اختبر معارفك في علم الطفيليات!","en":"Test your parasitology knowledge!"}.get(st.session_state.lang,"") }</h3>
        <p style='opacity:.5;'>{ {"fr":f"{len(QUIZ_QUESTIONS)} questions disponibles","ar":f"{len(QUIZ_QUESTIONS)} سؤال متاح","en":f"{len(QUIZ_QUESTIONS)} questions available"}.get(st.session_state.lang,"") }</p>
        </div>""", unsafe_allow_html=True)

        st.markdown("---")
        qc1, qc2 = st.columns(2)
        with qc1:
            n_questions = st.slider(
                {"fr": "Nombre de questions:", "ar": "عدد الأسئلة:", "en": "Number of questions:"}.get(st.session_state.lang, "Questions:"),
                5, min(25, len(QUIZ_QUESTIONS)), 10
            )
        with qc2:
            cats = list(set(q.get("cat", "General") for q in QUIZ_QUESTIONS))
            all_cat_label = {"fr": "Toutes les categories", "ar": "جميع الفئات", "en": "All categories"}.get(st.session_state.lang, "All")
            cats.insert(0, all_cat_label)
            chosen_cat = st.selectbox(
                {"fr": "Categorie:", "ar": "الفئة:", "en": "Category:"}.get(st.session_state.lang, "Category:"),
                cats
            )

        if st.button(f"🎮 {t('start_quiz')}", use_container_width=True, type="primary"):
            if chosen_cat == all_cat_label:
                pool = list(range(len(QUIZ_QUESTIONS)))
            else:
                pool = [i for i, q in enumerate(QUIZ_QUESTIONS) if q.get("cat") == chosen_cat]

            if len(pool) == 0:
                pool = list(range(len(QUIZ_QUESTIONS)))

            random.shuffle(pool)
            final_order = pool[:min(n_questions, len(pool))]

            st.session_state.quiz_state = {
                "current": 0,
                "score": 0,
                "answered": [],
                "active": True,
                "order": final_order,
                "wrong": [],
                "total_q": len(final_order),
                "finished": False,
                "selected_answer": None,
                "show_result": False
            }
            db_log(st.session_state.user_id, st.session_state.user_name, "Quiz started", f"n={len(final_order)} cat={chosen_cat}")
            st.rerun()

    # Quiz active - answering questions
    elif qs.get("active", False) and not qs.get("finished", False):
        idx = qs["current"]
        order = qs.get("order", [])
        total_q = qs.get("total_q", len(order))

        if idx < len(order):
            qi = order[idx]
            q = QUIZ_QUESTIONS[qi]

            # Progress
            progress_val = idx / total_q if total_q > 0 else 0
            st.progress(progress_val)

            q_num_label = {"fr": "Question", "ar": "سؤال", "en": "Question"}.get(st.session_state.lang, "Question")
            st.markdown(f"### {q_num_label} {idx + 1}/{total_q}")

            cat = q.get("cat", "")
            if cat:
                st.caption(f"📂 {cat}")

            q_text = tl(q["q"])
            st.markdown(f"""<div class='dm-card dm-card-purple'>
            <h4 style='margin:0;line-height:1.6;'>{q_text}</h4>
            </div>""", unsafe_allow_html=True)

            # Check if already answered this question
            if not qs.get("show_result", False):
                st.markdown("---")
                option_cols = st.columns(2)
                for i, opt in enumerate(q["opts"]):
                    with option_cols[i % 2]:
                        letter = ['A', 'B', 'C', 'D'][i]
                        if st.button(f"{letter}. {opt}", key=f"quiz_opt_{idx}_{i}", use_container_width=True):
                            correct = (i == q["ans"])
                            st.session_state.quiz_state["selected_answer"] = i
                            st.session_state.quiz_state["show_result"] = True
                            if correct:
                                st.session_state.quiz_state["score"] += 1
                            else:
                                st.session_state.quiz_state["wrong"].append({
                                    "q": q_text,
                                    "your": opt,
                                    "correct": q["opts"][q["ans"]]
                                })
                            st.session_state.quiz_state["answered"].append(correct)
                            st.rerun()
            else:
                # Show result of answer
                selected = qs.get("selected_answer", -1)
                correct_idx = q["ans"]
                is_correct = selected == correct_idx

                if is_correct:
                    st.success(f"✅ { {'fr':'Bonne reponse !','ar':'إجابة صحيحة!','en':'Correct!'}.get(st.session_state.lang,'Correct!') }")
                else:
                    correct_ans = q["opts"][correct_idx]
                    st.error(f"❌ { {'fr':'Reponse correcte','ar':'الإجابة الصحيحة','en':'Correct answer'}.get(st.session_state.lang,'Correct answer') }: **{correct_ans}**")

                # Show explanation
                expl = tl(q.get("expl", {}))
                if expl:
                    st.info(f"📖 {expl}")

                # Show all options with markers
                for i, opt in enumerate(q["opts"]):
                    if i == correct_idx:
                        st.markdown(f"✅ **{['A','B','C','D'][i]}. {opt}**")
                    elif i == selected and not is_correct:
                        st.markdown(f"❌ ~~{['A','B','C','D'][i]}. {opt}~~")
                    else:
                        st.markdown(f"{'  '}{['A','B','C','D'][i]}. {opt}")

                st.markdown("---")

                # Next question button
                if idx + 1 < len(order):
                    if st.button(f"➡️ {t('next_question')}", use_container_width=True, type="primary"):
                        st.session_state.quiz_state["current"] += 1
                        st.session_state.quiz_state["show_result"] = False
                        st.session_state.quiz_state["selected_answer"] = None
                        st.rerun()
                else:
                    finish_label = {"fr": "🏁 Voir les resultats", "ar": "🏁 عرض النتائج", "en": "🏁 See Results"}.get(st.session_state.lang, "🏁 Results")
                    if st.button(finish_label, use_container_width=True, type="primary"):
                        st.session_state.quiz_state["finished"] = True
                        st.session_state.quiz_state["active"] = False
                        st.rerun()

        else:
            # Fallback: mark as finished
            st.session_state.quiz_state["finished"] = True
            st.session_state.quiz_state["active"] = False
            st.rerun()

    # Quiz finished - show results
    elif qs.get("finished", False):
        score = qs.get("score", 0)
        total_q = qs.get("total_q", 1)
        pct = int(score / total_q * 100) if total_q > 0 else 0

        if pct >= 80:
            emoji, msg = "🏆", t("score_excellent")
        elif pct >= 60:
            emoji, msg = "👍", t("score_good")
        elif pct >= 40:
            emoji, msg = "📚", t("score_average")
        else:
            emoji, msg = "💪", t("score_low")

        st.markdown(f"""<div class='dm-card dm-card-green' style='text-align:center;'>
        <div style='font-size:5rem;'>{emoji}</div>
        <h2 class='dm-nt'>{t('result')}</h2>
        <div style='font-size:4rem;font-weight:900;font-family:JetBrains Mono,monospace;
            background:linear-gradient(135deg,#00f5ff,#ff00ff,#00ff88);
            -webkit-background-clip:text;-webkit-text-fill-color:transparent;'>
            {score}/{total_q}
        </div>
        <p style='font-size:1.5rem;opacity:.8;'>{pct}%</p>
        <p style='font-size:1.1rem;margin-top:10px;'>{msg}</p>
        </div>""", unsafe_allow_html=True)

        # Save score to DB
        try:
            db_quiz_save(st.session_state.user_id, st.session_state.user_name, score, total_q, pct)
            db_log(st.session_state.user_id, st.session_state.user_name, "Quiz done", f"{score}/{total_q}={pct}%")
        except Exception:
            pass

        # Performance chart
        if HAS_PLOTLY and total_q > 0:
            st.markdown("---")
            analysis_label = {"fr": "Analyse des resultats", "ar": "تحليل النتائج", "en": "Results Analysis"}.get(st.session_state.lang, "Analysis")
            st.markdown(f"### 📊 {analysis_label}")
            correct_label = {"fr": "Correctes", "ar": "صحيحة", "en": "Correct"}.get(st.session_state.lang, "Correct")
            incorrect_label = {"fr": "Incorrectes", "ar": "خاطئة", "en": "Incorrect"}.get(st.session_state.lang, "Incorrect")
            fig = go.Figure(data=[go.Pie(
                labels=[correct_label, incorrect_label],
                values=[score, total_q - score],
                marker_colors=["#00ff88", "#ff0040"],
                hole=0.5,
                textinfo='label+percent',
                textfont_size=14
            )])
            fig.update_layout(height=280, template=plot_template, margin=dict(l=20, r=20, t=20, b=20))
            st.plotly_chart(fig, use_container_width=True)

        # Wrong answers review
        wrong = qs.get("wrong", [])
        if wrong:
            review_label = {"fr": f"Erreurs a revoir ({len(wrong)})", "ar": f"الأخطاء ({len(wrong)})", "en": f"Mistakes to review ({len(wrong)})"}.get(st.session_state.lang, f"Mistakes ({len(wrong)})")
            with st.expander(f"❌ {review_label}"):
                for i, w in enumerate(wrong):
                    your_label = {"fr": "Votre reponse", "ar": "إجابتك", "en": "Your answer"}.get(st.session_state.lang, "Your answer")
                    correct_label2 = {"fr": "Correcte", "ar": "الصحيحة", "en": "Correct"}.get(st.session_state.lang, "Correct")
                    st.markdown(f"""**{i + 1}. {w['q']}**
- ❌ {your_label}: ~~{w['your']}~~
- ✅ {correct_label2}: **{w['correct']}**
---""")

        st.markdown("---")
        if st.button(f"🔄 {t('restart')}", use_container_width=True, type="primary"):
            st.session_state.quiz_state = {
                "current": 0, "score": 0, "answered": [],
                "active": False, "order": [], "wrong": [],
                "total_q": 0, "finished": False, "selected_answer": None,
                "show_result": False
            }
            st.rerun()


# ════════════════════════════════════════════
#  PAGE: CHATBOT (Fixed & Enhanced)
# ════════════════════════════════════════════
elif pg == "chat":
    st.title(f"💬 DM Bot")
    st.markdown(f"""<div class='dm-card dm-card-cyan'>
    <div style='display:flex;align-items:center;gap:15px;'>
        <div style='font-size:2.5rem;'>🤖</div>
        <div>
            <h4 style='margin:0;' class='dm-nt'>DM Bot</h4>
            <p style='margin:0;opacity:.5;font-size:.85rem;'>{ {"fr":"Assistant medical intelligent specialise en parasitologie",
    "ar":"مساعد طبي ذكي متخصص في علم الطفيليات",
    "en":"Intelligent medical assistant specialized in parasitology"}.get(st.session_state.lang,"")}</p>
        </div>
    </div>
    </div>""", unsafe_allow_html=True)

    # Initialize chat
    if "chat_history" not in st.session_state:
        st.session_state.chat_history = []

    if not st.session_state.chat_history:
        st.session_state.chat_history.append({"role": "bot", "msg": t("chat_welcome")})

    # Display chat history
    chat_container = st.container()
    with chat_container:
        for msg_item in st.session_state.chat_history:
            if msg_item["role"] == "user":
                st.markdown(f"""<div style='display:flex;justify-content:flex-end;margin:8px 0;'>
                <div class='dm-ch dm-cu'>👤 {msg_item['msg'] }</div>
                </div>""", unsafe_allow_html=True)
            else:
                st.markdown(f"""<div style='display:flex;justify-content:flex-start;margin:8px 0;'>
                <div class='dm-ch dm-cb'>🤖 {msg_item['msg'] }</div>
                </div>""", unsafe_allow_html=True)

    st.markdown("---")

    # Input using form to prevent issues
    with st.form("chat_form", clear_on_submit=True):
        user_input = st.text_input(t("chat_placeholder"), key="chat_input_field", label_visibility="collapsed",
                                    placeholder=t("chat_placeholder"))
        col_send, col_clear = st.columns([4, 1])
        with col_send:
            send_btn = st.form_submit_button(
                {"fr": "📨 Envoyer", "ar": "📨 إرسال", "en": "📨 Send"}.get(st.session_state.lang, "📨 Send"),
                use_container_width=True
            )
        with col_clear:
            clear_btn = st.form_submit_button(
                f"🗑️ {t('clear_chat')}",
                use_container_width=True
            )

    if send_btn and user_input and user_input.strip():
        st.session_state.chat_history.append( {"role": "user", "msg": user_input.strip()} )
        reply = chatbot_reply(user_input.strip())
        st.session_state.chat_history.append({"role": "bot", "msg": reply})
        db_log(st.session_state.user_id, st.session_state.user_name, "Chat", user_input[:100])
        st.rerun()

    if clear_btn:
        st.session_state.chat_history = []
        st.rerun()

    # Quick questions
    st.markdown(f"**{t('quick_questions')}**")

    # Row 1 - Parasites
    qr1_items = ["Amoeba", "Giardia", "Plasmodium", "Leishmania", "Trypanosoma", "Schistosoma", "Toxoplasma"]
    qr1_cols = st.columns(len(qr1_items))
    for col, q_item in zip(qr1_cols, qr1_items):
        with col:
            if st.button(q_item, key=f"cq1_{q_item}", use_container_width=True):
                st.session_state.chat_history.append({"role": "user", "msg": q_item})
                st.session_state.chat_history.append({"role": "bot", "msg": chatbot_reply(q_item)})
                st.rerun()

    # Row 2 - More
    qr2_items = ["Ascaris", "Taenia", "Oxyure", "Cryptosporidium", "Microscope", "Coloration", "Concentration"]
    qr2_cols = st.columns(len(qr2_items))
    for col, q_item in zip(qr2_cols, qr2_items):
        with col:
            if st.button(q_item, key=f"cq2_{q_item}", use_container_width=True):
                st.session_state.chat_history.append({"role": "user", "msg": q_item})
                st.session_state.chat_history.append({"role": "bot", "msg": chatbot_reply(q_item)})
                st.rerun()

    # Row 3 - Topics
    q3_labels_map = {
        "fr": ["Traitement", "Hygiene", "Selle (EPS)", "Aide"],
        "ar": ["علاج", "نظافة", "فحص البراز", "مساعدة"],
        "en": ["Treatment", "Hygiene", "Stool Exam", "Help"]
    }
    q3_keys_map = ["traitement", "hygiene", "selle", "aide"]
    q3_labels = q3_labels_map.get(st.session_state.lang, q3_labels_map["fr"])

    qr3_cols = st.columns(len(q3_labels))
    for col, (label, key) in zip(qr3_cols, zip(q3_labels, q3_keys_map)):
        with col:
            if st.button(label, key=f"cq3_{key}", use_container_width=True):
                st.session_state.chat_history.append({"role": "user", "msg": label})
                st.session_state.chat_history.append({"role": "bot", "msg": chatbot_reply(key)})
                st.rerun()


# ════════════════════════════════════════════
#  PAGE: COMPARISON (Enhanced)
# ════════════════════════════════════════════
elif pg == "cmp":
    st.title(f"🔄 {t('compare')}")

    desc = {"fr": "Comparez deux images microscopiques avec analyse avancee",
            "ar": "قارن بين صورتين مجهريتين بتحليل متقدم",
            "en": "Compare two microscopic images with advanced analysis"}.get(st.session_state.lang, "")

    st.markdown(f"""<div class='dm-card dm-card-cyan'>
    <p style='margin:0;'>{desc}</p>
    </div>""", unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        st.markdown(f"### 📷 {t('image1')}")
        f1 = st.file_uploader("img1", type=["jpg", "jpeg", "png", "bmp"], key="cmp1", label_visibility="collapsed")
    with c2:
        st.markdown(f"### 📷 {t('image2')}")
        f2 = st.file_uploader("img2", type=["jpg", "jpeg", "png", "bmp"], key="cmp2", label_visibility="collapsed")

    if f1 and f2:
        i1 = Image.open(f1).convert("RGB")
        i2 = Image.open(f2).convert("RGB")

        c1, c2 = st.columns(2)
        with c1:
            st.image(i1, caption=t("image1"), use_container_width=True)
        with c2:
            st.image(i2, caption=t("image2"), use_container_width=True)

        st.markdown("---")
        if st.button(f"🔍 {t('compare_btn')}", use_container_width=True, type="primary"):
            with st.spinner("Analyzing..."):
                metrics = compare_imgs(i1, i2)

            # Metrics display
            compare_result_label = {"fr": "Resultats de la comparaison", "ar": "نتائج المقارنة", "en": "Comparison Results"}.get(st.session_state.lang, "Results")
            st.markdown(f"### 📊 {compare_result_label}")
            mc = st.columns(4)
            with mc[0]:
                st.markdown(f"""<div class='dm-m'><span class='dm-m-i'>📊</span>
                <div class='dm-m-v'>{metrics['sim']}%</div>
                <div class='dm-m-l'>{t('similarity')}</div></div>""", unsafe_allow_html=True)
            with mc[1]:
                st.markdown(f"""<div class='dm-m'><span class='dm-m-i'>🎯</span>
                <div class='dm-m-v'>{metrics['ssim']}</div>
                <div class='dm-m-l'>SSIM</div></div>""", unsafe_allow_html=True)
            with mc[2]:
                st.markdown(f"""<div class='dm-m'><span class='dm-m-i'>📐</span>
                <div class='dm-m-v'>{metrics['mse']}</div>
                <div class='dm-m-l'>MSE</div></div>""", unsafe_allow_html=True)
            with mc[3]:
                if metrics["sim"] > 90:
                    verdict = {"fr": "Tres similaires", "ar": "متشابهتان جدا", "en": "Very similar"}
                    v_icon = "✅"
                elif metrics["sim"] > 70:
                    verdict = {"fr": "Similaires", "ar": "متشابهتان", "en": "Similar"}
                    v_icon = "🟡"
                elif metrics["sim"] > 50:
                    verdict = {"fr": "Peu similaires", "ar": "قليل التشابه", "en": "Somewhat similar"}
                    v_icon = "🟠"
                else:
                    verdict = {"fr": "Tres differentes", "ar": "مختلفتان جدا", "en": "Very different"}
                    v_icon = "🔴"
                st.markdown(f"""<div class='dm-m'><span class='dm-m-i'>🔍</span>
                <div class='dm-m-v' style='font-size:1rem;'>{v_icon} {tl(verdict)}</div>
                <div class='dm-m-l'>Verdict</div></div>""", unsafe_allow_html=True)

            # SSIM Gauge
            if HAS_PLOTLY:
                st.markdown("---")
                fig = go.Figure(go.Indicator(
                    mode="gauge+number",
                    value=metrics["sim"],
                    title={"text": t("similarity"), "font": {"color": "#e0e8ff"}},
                    number={"font": {"color": "#00f5ff"}},
                    gauge={
                        "axis": {"range": [0, 100], "tickcolor": "#6b7fa0"},
                        "bar": {"color": "#00f5ff"},
                        "bgcolor": "rgba(10,15,46,0.5)",
                        "steps": [
                            {"range": [0, 30], "color": "rgba(255,0,64,0.3)"},
                            {"range": [30, 60], "color": "rgba(255,149,0,0.3)"},
                            {"range": [60, 80], "color": "rgba(255,255,0,0.3)"},
                            {"range": [80, 100], "color": "rgba(0,255,136,0.3)"}
                        ],
                        "threshold": {"line": {"color": "white", "width": 4}, "thickness": 0.75, "value": metrics["sim"]}
                    }
                ))
                fig.update_layout(height=280, template=plot_template,
                                  margin=dict(l=30, r=30, t=60, b=20),
                                  paper_bgcolor="rgba(0,0,0,0)",
                                  font={"color": "#e0e8ff"})
                st.plotly_chart(fig, use_container_width=True)

            # Pixel difference
            st.markdown(f"### 🔍 {t('pixel_diff')}")
            diff_img = pixel_diff(i1, i2)
            dc1, dc2, dc3 = st.columns(3)
            with dc1:
                st.image(i1, caption=t("image1"), use_container_width=True)
            with dc2:
                st.image(diff_img, caption=t("pixel_diff"), use_container_width=True)
            with dc3:
                st.image(i2, caption=t("image2"), use_container_width=True)

            # Filter comparison
            st.markdown(f"### 🔬 {t('filter_comparison')}")
            filter_list = [
                ({"fr": "Thermique", "ar": "حراري", "en": "Thermal"}, thermal),
                ({"fr": "Contours", "ar": "حواف", "en": "Edges"}, edges_filter),
                ({"fr": "Contraste+", "ar": "تباين+", "en": "Enhanced"}, enhanced_filter),
                ({"fr": "Negatif", "ar": "سلبي", "en": "Negative"}, negative_filter),
                ({"fr": "Relief", "ar": "نقش", "en": "Emboss"}, emboss_filter),
            ]
            for fname, ffunc in filter_list:
                fn = tl(fname)
                fc1, fc2 = st.columns(2)
                with fc1:
                    st.image(ffunc(i1), caption=f"{t('image1')} - {fn}", use_container_width=True)
                with fc2:
                    st.image(ffunc(i2), caption=f"{t('image2')} - {fn}", use_container_width=True)

            # Histogram comparison
            if HAS_PLOTLY:
                hist_label = {"fr": "Comparaison des histogrammes", "ar": "مقارنة المدرجات التكرارية", "en": "Histogram Comparison"}.get(st.session_state.lang, "Histogram")
                st.markdown(f"### 📊 {hist_label}")
                h1 = histogram(i1)
                h2 = histogram(i2)
                hc1, hc2 = st.columns(2)
                with hc1:
                    fig = go.Figure()
                    fig.add_trace(go.Scatter(y=h1["red"], name="R", line=dict(color="red", width=1)))
                    fig.add_trace(go.Scatter(y=h1["green"], name="G", line=dict(color="green", width=1)))
                    fig.add_trace(go.Scatter(y=h1["blue"], name="B", line=dict(color="blue", width=1)))
                    fig.update_layout(title=t("image1"), height=250, template=plot_template, margin=dict(l=20, r=20, t=40, b=20))
                    st.plotly_chart(fig, use_container_width=True)
                with hc2:
                    fig = go.Figure()
                    fig.add_trace(go.Scatter(y=h2["red"], name="R", line=dict(color="red", width=1)))
                    fig.add_trace(go.Scatter(y=h2["green"], name="G", line=dict(color="green", width=1)))
                    fig.add_trace(go.Scatter(y=h2["blue"], name="B", line=dict(color="blue", width=1)))
                    fig.update_layout(title=t("image2"), height=250, template=plot_template, margin=dict(l=20, r=20, t=40, b=20))
                    st.plotly_chart(fig, use_container_width=True)


# ════════════════════════════════════════════
#  PAGE: ADMIN
# ════════════════════════════════════════════
elif pg == "admin":
    st.title(f"⚙️ {t('admin')}")
    if not has_role(3):
        st.error("🔒 Admin access required")
        st.stop()

    tab1, tab2, tab3 = st.tabs([f"👥 {t('users_mgmt')}", f"📜 {t('activity_log')}", f"🖥️ {t('system_info')}"])

    with tab1:
        users = db_users()
        if users:
            st.dataframe(pd.DataFrame(users), use_container_width=True)
            st.markdown("---")

            tc1, tc2, tc3 = st.columns(3)
            uid = tc1.number_input("User ID", min_value=1, step=1, key="toggle_uid")
            act = tc2.selectbox("Status", ["Active", "Disabled"], key="toggle_status")
            if tc3.button("Apply", use_container_width=True, key="toggle_btn"):
                db_toggle(uid, act == "Active")
                db_log(st.session_state.user_id, st.session_state.user_name, "Toggle user", f"#{uid}={act}")
                st.success(f"User #{uid} -> {act}")
                st.rerun()

        st.markdown("---")
        st.markdown(f"### ➕ {t('create_user')}")
        with st.form("new_user_form"):
            nu = st.text_input(f"{t('username')} *", key="new_username")
            np2 = st.text_input(f"{t('password')} *", type="password", key="new_password")
            nf = st.text_input("Full Name *", key="new_fullname")
            nr = st.selectbox("Role", list(ROLES.keys()), key="new_role")
            ns = st.text_input("Speciality", "Laboratoire", key="new_spec")
            if st.form_submit_button(t("create_user"), use_container_width=True):
                if nu and np2 and nf:
                    if db_create_user(nu, np2, nf, nr, ns):
                        db_log(st.session_state.user_id, st.session_state.user_name, "Created user", nu)
                        st.success(f"User '{nu}' created!")
                        st.rerun()
                    else:
                        st.error("Username already exists")
                else:
                    st.error("Please fill all required fields")

        st.markdown("---")
        st.markdown(f"### 🔑 {t('change_pwd')}")
        cp_col1, cp_col2 = st.columns(2)
        cpid = cp_col1.number_input("User ID", min_value=1, step=1, key="chpw_uid")
        cpnew = cp_col2.text_input("New Password", type="password", key="chpw_new")
        if st.button(t("change_pwd"), key="chpw_btn"):
            if cpnew:
                db_chpw(cpid, cpnew)
                db_log(st.session_state.user_id, st.session_state.user_name, "Changed pwd", f"#{cpid}")
                st.success(f"Password updated for user #{cpid}")

    with tab2:
        logs = db_logs(300)
        if logs:
            ldf = pd.DataFrame(logs)
            if "username" in ldf.columns:
                usernames_list = sorted(ldf["username"].dropna().unique().tolist())
                flt = st.selectbox("Filter by user:", ["All"] + usernames_list, key="log_filter")
                if flt != "All":
                    ldf = ldf[ldf["username"] == flt]
            st.dataframe(ldf, use_container_width=True)
        else:
            st.info(t("no_data"))

    with tab3:
        sc1, sc2, sc3 = st.columns(3)
        with sc1:
            st.markdown(f"""<div class='dm-card dm-card-green'>
            <h4>🟢 System OK</h4>
            <p><b>Version:</b> {APP_VERSION}</p>
            <p><b>Python:</b> {os.sys.version.split()[0]}</p>
            <p><b>Bcrypt:</b> {'✅' if HAS_BCRYPT else '❌ (SHA256)'}</p>
            <p><b>Plotly:</b> {'✅' if HAS_PLOTLY else '❌'}</p>
            <p><b>QR Code:</b> {'✅' if HAS_QRCODE else '❌'}</p>
            <p><b>Database:</b> SQLite</p>
            </div>""", unsafe_allow_html=True)
        with sc2:
            ts = db_stats()
            st.markdown(f"""<div class='dm-card dm-card-cyan'>
            <h4>📊 Statistics</h4>
            <p><b>Users:</b> {len(db_users())}</p>
            <p><b>Analyses:</b> {ts['total']}</p>
            <p><b>Reliable:</b> {ts['reliable']}</p>
            <p><b>Quiz Scores:</b> {len(db_leaderboard())}</p>
            <p><b>Languages:</b> FR / AR / EN</p>
            </div>""", unsafe_allow_html=True)
        with sc3:
            dbsz = os.path.getsize(DB_PATH) / 1024 if os.path.exists(DB_PATH) else 0
            st.markdown(f"""<div class='dm-card'>
            <h4>💾 Storage</h4>
            <p><b>DB Size:</b> {dbsz:.1f} KB</p>
            <p><b>Parasites:</b> {len(CLASS_NAMES)} classes</p>
            <p><b>Quiz:</b> {len(QUIZ_QUESTIONS)} questions</p>
            <p><b>Chat KB:</b> {len(CHAT_KB)} entries</p>
            </div>""", unsafe_allow_html=True)


# ════════════════════════════════════════════
#  PAGE: ABOUT
# ════════════════════════════════════════════
elif pg == "about":
    st.title(f"ℹ️ {t('about')}")
    lang = st.session_state.lang

    st.markdown(f"""<div class='dm-card dm-card-cyan' style='text-align:center;'>
    <h1 class='dm-nt'>🧬 DM SMART LAB AI</h1>
    <p style='font-size:1.1rem;font-family:Orbitron,sans-serif;'><b>v{APP_VERSION} — Space Edition</b></p>
    <p style='opacity:.4;'>{t('system_desc')}</p>
    </div>""", unsafe_allow_html=True)

    desc_about = {
        "fr": "Ce projet innovant utilise les technologies de Deep Learning et de Vision par Ordinateur pour assister les techniciens de laboratoire dans l'identification rapide et precise des parasites.",
        "ar": "يستخدم هذا المشروع المبتكر تقنيات التعلم العميق والرؤية الحاسوبية لمساعدة تقنيي المخبر في التعرف السريع والدقيق على الطفيليات.",
        "en": "This innovative project uses Deep Learning and Computer Vision technologies to assist laboratory technicians in the rapid and accurate identification of parasites."
    }.get(lang, "")

    st.markdown(f"""<div class='dm-card'>
    <h3>📖 {tl(PROJECT_TITLE)}</h3>
    <p style='line-height:1.8;opacity:.8;'>{desc_about}</p>
    </div>""", unsafe_allow_html=True)

    st.markdown("---")
    c1, c2 = st.columns(2)
    with c1:
        d1r = tl(AUTHORS['dev1']['role'])
        d2r = tl(AUTHORS['dev2']['role'])
        level_label = {"fr": "Niveau", "ar": "المستوى", "en": "Level"}.get(lang, "Level")
        year_label = {"fr": "3eme Annee", "ar": "السنة الثالثة", "en": "3rd Year"}.get(lang, "3rd Year")
        spec_label = {"fr": "Specialite", "ar": "التخصص", "en": "Speciality"}.get(lang, "Speciality")
        spec_val = {"fr": "Laboratoire de Sante Publique", "ar": "مخبر الصحة العمومية", "en": "Public Health Laboratory"}.get(lang, "Public Health Lab")

        st.markdown(f"""<div class='dm-card dm-card-cyan'>
        <h3>👨‍🔬 {t('dev_team')}</h3><br>
        <p><b>🧑‍💻 {AUTHORS['dev1']['name']}</b><br><span style='opacity:.5;'>{d1r}</span></p><br>
        <p><b>🔬 {AUTHORS['dev2']['name']}</b><br><span style='opacity:.5;'>{d2r}</span></p><br>
        <p><b>{level_label}:</b> {year_label}</p>
        <p><b>{spec_label}:</b> {spec_val}</p>
        </div>""", unsafe_allow_html=True)
    with c2:
        obj_label = {"fr": "Objectifs", "ar": "الأهداف", "en": "Objectives"}.get(lang, "Objectives")
        obj1 = {"fr": "Automatiser la lecture microscopique", "ar": "أتمتة القراءة المجهرية", "en": "Automate microscopic reading"}.get(lang, "")
        obj2 = {"fr": "Reduire les erreurs diagnostiques", "ar": "تقليل الأخطاء التشخيصية", "en": "Reduce diagnostic errors"}.get(lang, "")
        obj3 = {"fr": "Accelerer le processus d'analyse", "ar": "تسريع عملية التحليل", "en": "Speed up analysis process"}.get(lang, "")
        obj4 = {"fr": "Assister les professionnels de sante", "ar": "مساعدة المهنيين الصحيين", "en": "Assist healthcare professionals"}.get(lang, "")

        st.markdown(f"""<div class='dm-card'>
        <h3>🏫 {t('institution')}</h3><br>
        <p><b>{tl(INSTITUTION['name'])}</b></p>
        <p>📍 {INSTITUTION['city']}, {tl(INSTITUTION['country'])} 🇩🇿</p><br>
        <h4>🎯 {obj_label}</h4>
        <ul>
        <li>{obj1}</li>
        <li>{obj2}</li>
        <li>{obj3}</li>
        <li>{obj4}</li>
        </ul>
        </div>""", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown(f"### 🛠️ {t('technologies')}")
    tc = st.columns(8)
    techs = [("🐍", "Python", "Core"), ("🧠", "TensorFlow", "AI"), ("🎨", "Streamlit", "UI"),
             ("📊", "Plotly", "Charts"), ("🗄️", "SQLite", "DB"), ("🔒", "Bcrypt", "Security"),
             ("📄", "FPDF", "PDF"), ("📱", "QR", "Verify")]
    for col, (ic, n, d) in zip(tc, techs):
        with col:
            st.markdown(f"""<div class='dm-m'>
            <span class='dm-m-i'>{ic}</span>
            <div class='dm-m-v' style='font-size:.85rem;'>{n}</div>
            <div class='dm-m-l'>{d}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown(f"### 🌟 Features v{APP_VERSION}")
    feat_cols = st.columns(4)
    features = [
        ("📸", "Camera", "Direct capture"),
        ("🌍", "3 Languages", "FR/AR/EN"),
        ("🤖", "DM Bot", "AI Chatbot"),
        ("🧠", "30+ Quiz", "By category"),
        ("🔬", "7 Parasites", "Complete DB"),
        ("🎯", "Heatmap", "AI Visualization"),
        ("📄", "PDF Pro", "QR + Signatures"),
        ("🔄", "Compare+", "Pixel diff + filters"),
        ("🔊", "Voice", "Web Speech API"),
        ("📊", "Plotly", "Pro charts"),
        ("🔐", "Secure Auth", "Bcrypt + lockout"),
        ("🌌", "Space Theme", "Animated dark UI"),
    ]
    for i, (ic, name, desc) in enumerate(features):
        with feat_cols[i % 4]:
            st.markdown(f"""<div class='dm-card' style='padding:14px;text-align:center;'>
            <div style='font-size:1.8rem;'>{ic}</div>
            <p style='font-weight:700;margin:4px 0;font-size:.85rem;'>{name}</p>
            <p style='font-size:.7rem;opacity:.4;'>{desc}</p>
            </div>""", unsafe_allow_html=True)

    st.markdown("---")
    made_label = {"fr": "Fait avec", "ar": "صنع بـ", "en": "Made with"}.get(lang, "Made with")
    in_label = {"fr": "a", "ar": "في", "en": "in"}.get(lang, "in")
    st.caption(f"{made_label} ❤️ {in_label} {INSTITUTION['city']} — {INSTITUTION['year']} 🇩🇿")     
